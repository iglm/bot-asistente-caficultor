"""
Tests para la exportación de Excel.

Verifica que ExcelManager.generar_excel() produce un archivo válido
con las hojas correctas. Usa DB temporal para evitar tocar la BD real.
"""

import os
from pathlib import Path

import pytest
from database import Database
from excel_manager import ExcelManager

# Ruta al template Excel real
TEMPLATE_PATH = os.path.join(
    Path(__file__).resolve().parent.parent,
    "data", "plantilla", "Costos de produccion - 2026.xlsx",
)

# Hojas esperadas en el Excel generado después de eliminar vacías
# (basado en la lógica de ExcelManager.generar_excel)
HOJAS_ESPERADAS = [
    "ID lotes",
    "Ingresos",
    "Instalacion de Cafe",
    "Control de arvenses",
    "Fertilizacion",
    "Control Fitosanitario",
    "Regulacion de sombrio",
    "Otras Labores",
    "Recoleccion",
    "Beneficio",
    "Gastos Administrativos",
    "Gráficos",
    "Resumen Ejecutivo",
    "Dashboard",
    "Configuración",
]


@pytest.fixture
def excel_manager():
    """Instancia de ExcelManager apuntando al template real."""
    if not os.path.exists(TEMPLATE_PATH):
        pytest.skip(f"Template Excel no encontrado: {TEMPLATE_PATH}")
    return ExcelManager(template_path=TEMPLATE_PATH)


@pytest.fixture
def db_con_datos(tmp_path):
    """DB temporal con una finca, lotes y transacciones para exportar."""
    db_path = str(tmp_path / "test_export.db")
    db = Database(db_path=db_path)
    db.init_db()

    user_id = 1001
    db.upsert_user(user_id=user_id, username="export_user")
    finca_id = db.create_finca(
        user_id=user_id,
        nombre="Finca Export Test",
        region="Caldas",
        departamento="Manizales",
    )

    # Lotes
    db.create_lote(finca_id, "Lote Export A", area=5.0, num_arboles=5000,
                   variedad="Castillo", fecha_siembra="2022-01-01")
    db.create_lote(finca_id, "Lote Export B", area=3.0, num_arboles=3000,
                   variedad="Caturra", fecha_siembra="2021-06-01")

    # Transacciones
    db.insert_transaccion(finca_id, "ingreso_cps", "2024-10-15",
                          producto="CPS", cantidad=2000, valor_unitario=18000,
                          valor_total=36_000_000)
    db.insert_transaccion(finca_id, "ingreso_pasilla", "2024-10-15",
                          producto="Pasilla", cantidad=200, valor_unitario=3000,
                          valor_total=600_000)
    db.insert_transaccion(finca_id, "recoleccion", "2024-10-01",
                          labor="Corte", cantidad=50, unidad="jornal",
                          valor_unitario=55000, valor_total=2_750_000)
    db.insert_transaccion(finca_id, "fertilizacion_mo", "2024-03-10",
                          labor="Aplicación", cantidad=10, unidad="jornal",
                          valor_unitario=55000, valor_total=550_000)
    db.insert_transaccion(finca_id, "fertilizacion_insumos", "2024-03-10",
                          producto="Urea", cantidad=200, unidad="kg",
                          valor_unitario=3200, valor_total=640_000)

    return db, finca_id


class TestExportacionExcel:
    """Prueba la generación de archivos Excel."""

    def test_excel_se_genera(self, excel_manager, db_con_datos, tmp_path):
        """El archivo Excel se genera correctamente en la ruta indicada."""
        db, finca_id = db_con_datos
        output_path = str(tmp_path / "test_export.xlsx")

        resultado = excel_manager.generar_excel(finca_id, db, output_path)

        assert resultado == output_path
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 0

    def test_excel_tiene_hojas_correctas(self, excel_manager, db_con_datos, tmp_path):
        """El Excel generado contiene las hojas de datos esperadas."""
        import openpyxl

        db, finca_id = db_con_datos
        output_path = str(tmp_path / "test_hojas.xlsx")

        excel_manager.generar_excel(finca_id, db, output_path)

        wb = openpyxl.load_workbook(output_path, read_only=True)
        hojas = wb.sheetnames
        wb.close()

        # Verificar que las hojas críticas existen
        for hoja in ["ID lotes", "Ingresos por ventas de cafe", "Recoleccion", "Gráficos",
                      "Resumen Ejecutivo", "Dashboard", "Configuración"]:
            assert hoja in hojas, f"Hoja '{hoja}' no encontrada en {hojas}"

        # Al menos las hojas de datos principales deben estar
        assert "Instalacion de Cafe" in hojas
        assert "Gastos Administrativos" in hojas

    def test_excel_sin_hojas_vacias(self, excel_manager, db_con_datos, tmp_path):
        """
        El Excel generado NO debe contener las hojas vacías que
        ExcelManager elimina explícitamente (Plan de ordenamiento, etc.).
        """
        import openpyxl

        db, finca_id = db_con_datos
        output_path = str(tmp_path / "test_sin_vacias.xlsx")

        excel_manager.generar_excel(finca_id, db, output_path)

        wb = openpyxl.load_workbook(output_path, read_only=True)
        hojas = wb.sheetnames
        wb.close()

        hojas_eliminadas = ["Plan de ordenamiento", "Plan de acción", "Cronograma"]
        for hoja in hojas_eliminadas:
            assert hoja not in hojas, f"Hoja vacía '{hoja}' no debería estar presente"

    def test_excel_contiene_datos_en_hoja_lotes(self, excel_manager, db_con_datos, tmp_path):
        """La hoja 'ID lotes' contiene filas de datos (no solo headers)."""
        import openpyxl

        db, finca_id = db_con_datos
        output_path = str(tmp_path / "test_lotes.xlsx")

        excel_manager.generar_excel(finca_id, db, output_path)

        wb = openpyxl.load_workbook(output_path, data_only=True)
        ws = wb["ID lotes"]
        filas_con_datos = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None for cell in row):
                filas_con_datos += 1
        wb.close()

        assert filas_con_datos >= 2, (
            f"Se esperaban al menos 2 lotes, se encontraron {filas_con_datos}"
        )

    def test_excel_sin_template_lanza_error(self, tmp_path):
        """Si el template no existe, ExcelManager debe lanzar FileNotFoundError."""
        manager = ExcelManager(template_path="/ruta/inexistente/template.xlsx")
        with pytest.raises(FileNotFoundError):
            manager._validar_template()

    def test_plantilla_vacia_se_genera(self, excel_manager, tmp_path):
        """generar_plantilla_vacia() produce un archivo válido."""
        output_path = str(tmp_path / "plantilla_vacia.xlsx")
        resultado = excel_manager.generar_plantilla_vacia(output_path)

        assert resultado == output_path
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 0

    def test_plantilla_vacia_tiene_hoja_notas(self, excel_manager, tmp_path):
        """La plantilla vacía debe incluir la hoja NOTAS con instrucciones."""
        import openpyxl

        output_path = str(tmp_path / "plantilla_notas.xlsx")
        excel_manager.generar_plantilla_vacia(output_path)

        wb = openpyxl.load_workbook(output_path, read_only=True)
        hojas = wb.sheetnames
        wb.close()

        assert "NOTAS" in hojas, "La plantilla vacía debe tener hoja NOTAS"
