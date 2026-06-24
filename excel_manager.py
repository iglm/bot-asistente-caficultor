"""
Manejador de Excel para el Bot Asistente Caficultor.
Carga el template, llena SOLO las hojas de datos con información de SQLite,
y guarda manteniendo fórmulas intactas.

VERSIÓN DINÁMICA: En lugar de extender fórmulas hasta límites fijos,
crea filas según los datos reales usando copy/paste de fórmulas.
"""
import os
import re
import shutil
import logging
from copy import copy
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Mapeo de categorías DB → hojas Excel y sus configuraciones
HOJA_CONFIG = {
    # (hoja, columnas_mo, columnas_insumos, columna_inicio_insumos)
    # None para columnas significa que esa sección no aplica
    "Instalacion de Cafe": {
        "mo_cols": {"start": "A", "end": "F", "campos": ["lote", "fecha", "labor", "cantidad", "valor_unitario", "valor_total_formula"]},
        "insumos_cols": {"start": "H", "end": "L", "campos": ["fecha", "producto", "cantidad", "valor_unitario", "valor_total_formula"]},
        "categorias_mo": ["instalacion_mo"],
        "categorias_insumos": ["instalacion_insumos"],
    },
    "Control de arvenses": {
        "mo_cols": {"start": "A", "end": "F", "campos": ["lote", "fecha", "labor", "cantidad", "valor_unitario", "valor_total_formula"]},
        "insumos_cols": {"start": "H", "end": "M", "campos": ["fecha", "labor", "producto", "cantidad", "valor_unitario", "valor_total_formula"]},
        "categorias_mo": ["arvenses_mo"],
        "categorias_insumos": ["arvenses_insumos"],
    },
    "Fertilizacion": {
        "mo_cols": {"start": "A", "end": "F", "campos": ["lote", "fecha", "labor", "cantidad", "valor_unitario", "valor_total_formula"]},
        "insumos_cols": {"start": "H", "end": "L", "campos": ["fecha", "producto", "cantidad", "valor_unitario", "valor_total_formula"]},
        "categorias_mo": ["fertilizacion_mo"],
        "categorias_insumos": ["fertilizacion_insumos"],
    },
    "Control Fitosanitario": {
        "mo_cols": {"start": "A", "end": "F", "campos": ["lote", "fecha", "labor", "cantidad", "valor_unitario", "valor_total_formula"]},
        "insumos_cols": {"start": "H", "end": "M", "campos": ["fecha", "labor", "producto", "cantidad", "valor_unitario", "valor_total_formula"]},
        "categorias_mo": ["fitosanitario_mo"],
        "categorias_insumos": ["fitosanitario_insumos"],
    },
    "Regulacion de sombrio": {
        "mo_cols": {"start": "A", "end": "F", "campos": ["lote", "fecha", "labor", "cantidad", "valor_unitario", "valor_total_formula"]},
        "insumos_cols": {"start": "H", "end": "L", "campos": ["fecha", "producto", "cantidad", "valor_unitario", "valor_total_formula"]},
        "categorias_mo": ["sombrio_mo"],
        "categorias_insumos": ["sombrio_insumos"],
    },
    "Otras Labores": {
        "mo_cols": {"start": "A", "end": "F", "campos": ["lote", "fecha", "labor", "cantidad", "valor_unitario", "valor_total_formula"]},
        "insumos_cols": {"start": "H", "end": "L", "campos": ["fecha", "producto", "cantidad", "valor_unitario", "valor_total_formula"]},
        "categorias_mo": ["otras_labores_mo"],
        "categorias_insumos": ["otras_labores_insumos"],
    },
    "Recoleccion": {
        # Recolección es especial: A=Fecha, B=Labor, C=Kilos, D=V.Unitario(formula=E/C), E=V.Total
        "columnas_especiales": ["fecha", "labor", "cantidad", "valor_unitario_formula", "valor_total"],
        "categorias": ["recoleccion"],
    },
    "Beneficio": {
        # Beneficio: A=Fecha, B=Labor, C=Jornales, D=V.Unitario, E=V.Total(formula=D*C)
        "columnas_especiales": ["fecha", "labor", "cantidad", "valor_unitario", "valor_total_formula"],
        "categorias": ["beneficio"],
    },
    "Gastos Administrativos": {
        # Gastos Admin: A=Fecha, B=Gasto, C=V.Total
        "columnas_especiales": ["fecha", "labor", "valor_total"],
        "categorias": ["administrativo"],
    },
}

# Hoja ID lotes tiene Tabla4 — no usar insert_rows que la rompe,
# mejor extender filas manualmente
HOJA_LOTES_START_ROW = 2
HOJA_LOTES_SUBTOTAL_ROW = 17
HOJA_LOTES_TEMPLATE_ROWS = 15  # filas 2-16

# Para hojas con datos desde fila 3 hasta subtotal en fila 20
HOJA_COSTOS_START_ROW = 3
HOJA_COSTOS_SUBTOTAL_ROW = 20
HOJA_COSTOS_TEMPLATE_ROWS = 17  # filas 3-19


class ExcelManager:
    """Manejador del template Excel. Solo llena datos, mantiene fórmulas."""

    def __init__(self, template_path: str):
        self.template_path = template_path

    def _validar_template(self):
        """Verifica que el template exista."""
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(
                f"Template Excel no encontrado: {self.template_path}"
            )

    # ------------------------------------------------------------------
    # Helpers para copiar fórmulas y formato (dinámicos)
    # ------------------------------------------------------------------

    @staticmethod
    def _ajustar_referencias_fila(formula: str, fila_origen: int, fila_destino: int) -> str:
        """
        Ajusta las referencias de celda en una fórmula de fila_origen a fila_destino.
        Ejemplo: '=IFERROR(F3/D3, 0)' con origen=3, destino=10 → '=IFERROR(F10/D10, 0)'
        Solo reemplaza referencias que apuntan EXACTAMENTE a fila_origen.
        """
        if fila_origen == fila_destino or not formula or not formula.startswith("="):
            return formula

        # Expresión regular: letras de columna (1-3) seguidas de dígitos de fila
        def _reemplazar(match):
            col = match.group(1)
            row = int(match.group(2))
            if row == fila_origen:
                return f"{col}{fila_destino}"
            return match.group(0)

        return re.sub(r'([A-Z]{1,3})(\d+)', _reemplazar, formula)

    @staticmethod
    def _copiar_estilo_celda(celda_origen, celda_destino):
        """Copia el formato (bordes, fuente, relleno, alineación, formato núm.) de una celda a otra."""
        if not celda_origen or not celda_destino:
            return
        try:
            if celda_origen.has_style:
                celda_destino.font = copy(celda_origen.font)
                celda_destino.border = copy(celda_origen.border)
                celda_destino.fill = copy(celda_origen.fill)
                celda_destino.number_format = copy(celda_origen.number_format)
                celda_destino.protection = copy(celda_origen.protection)
                celda_destino.alignment = copy(celda_origen.alignment)
        except Exception:
            pass  # Ignorar errores de copia de estilo (ej. fill inválido)

    @staticmethod
    def _copiar_formula_fila(ws, fila_origen: int, fila_destino: int, max_col: int):
        """
        Copia las fórmulas de fila_origen a fila_destino, ajustando referencias
        de celda. Preserva formato de todas las celdas.
        """
        for col in range(1, max_col + 1):
            celda_origen = ws.cell(row=fila_origen, column=col)
            celda_destino = ws.cell(row=fila_destino, column=col)

            # Copiar valor/fórmula ajustada
            val = celda_origen.value
            if val is not None:
                if isinstance(val, str) and val.startswith("="):
                    celda_destino.value = ExcelManager._ajustar_referencias_fila(
                        val, fila_origen, fila_destino
                    )
                elif isinstance(val, str):
                    # Copiar texto literal
                    celda_destino.value = val
                # No copiar valores numéricos/fechas — esos van con datos reales

            # Copiar formato
            ExcelManager._copiar_estilo_celda(celda_origen, celda_destino)

    def _asegurar_filas_suficientes(
        self, ws, data_start_row: int, subtotal_row_original: int,
        num_needed: int, template_rows: int, max_col: int
    ) -> tuple:
        """
        Asegura que haya suficientes filas de datos entre data_start_row y el subtotal.
        
        Args:
            ws: Worksheet
            data_start_row: Primera fila de datos (2 o 3)
            subtotal_row_original: Fila del subtotal en el template (17 o 20)
            num_needed: Número de filas de datos necesarias
            template_rows: Cuántas filas de datos tiene el template
            max_col: Número máximo de columnas a copiar
        
        Returns:
            tuple: (nuevo_subtotal_row, last_data_row)
        """
        if num_needed <= template_rows:
            # Suficientes filas en el template — solo devolver info
            last_data_row = data_start_row + num_needed - 1
            return subtotal_row_original, last_data_row

        # Necesitamos insertar filas
        rows_to_add = num_needed - template_rows
        last_existing_data = data_start_row + template_rows - 1  # ej: 19
        new_last_data = data_start_row + num_needed - 1  # ej: data_start + N - 1

        # Insertar filas justo después de la última fila de datos existente
        # (y antes del subtotal)
        ws.insert_rows(last_existing_data + 1, rows_to_add)

        # El subtotal se movió hacia abajo
        new_subtotal_row = subtotal_row_original + rows_to_add

        # Copiar fórmulas desde la primera fila de datos a las nuevas filas
        source_row = data_start_row
        for i in range(rows_to_add):
            target_row = last_existing_data + 1 + i
            self._copiar_formula_fila(ws, source_row, target_row, max_col)

        # Actualizar referencia de tabla Excel si existe (ID lotes)
        # ws.tables devuelve TableList (dict-like) con objetos Table como valores
        # En openpyxl 3.1.5, items() devuelve (name, Table) correctamente
        try:
            if ws.tables and "Tabla4" in ws.tables:
                table = ws.tables["Tabla4"]
                if hasattr(table, "ref"):
                    table_ref = table.ref
                    # ref format: A1:F17 → cambiar F17 a F(new_subtotal_row)
                    match = re.match(r'^([A-Z]+)(\d+):([A-Z]+)(\d+)$', table_ref)
                    if match:
                        col1, _, col2, _ = match.groups()
                        new_ref = f"{col1}{data_start_row-1}:{col2}{new_subtotal_row}"
                        table.ref = new_ref
                        logger.debug(f"Tabla {table.displayName} expandida: {table_ref} → {new_ref}")
        except Exception as e:
            logger.warning(f"No se pudo actualizar tabla Excel: {e}")

        return new_subtotal_row, new_last_data

    @staticmethod
    def _actualizar_sum_subtotal(ws, subtotal_row: int, data_start: int, data_end: int, celdas_sum: list):
        """
        Actualiza las fórmulas SUM en la fila de subtotal para cubrir el rango correcto.
        
        Args:
            ws: Worksheet
            subtotal_row: fila del subtotal
            data_start: primera fila de datos
            data_end: última fila de datos
            celdas_sum: lista de columnas (1-indexed) que tienen SUM
        """
        for col in celdas_sum:
            celda = ws.cell(row=subtotal_row, column=col)
            val = celda.value
            if val is not None and isinstance(val, str) and val.startswith("="):
                # Buscar SUM(rango) y actualizar el rango
                # Patrones: SUM(F3:F19), +SUM(C3:C19), etc.
                new_val = re.sub(
                    r'(SUM\()([A-Z]+)(\d+):([A-Z]+)(\d+)(\))',
                    lambda m: f"{m.group(1)}{m.group(2)}{data_start}:{m.group(4)}{data_end}{m.group(6)}",
                    val
                )
                if new_val != val:
                    celda.value = new_val
                    logger.debug(f"Subtotal col {col} actualizado: {val} → {new_val}")

    # ------------------------------------------------------------------
    # Generar plantilla vacía (sin datos, solo headers + ejemplo)
    # ------------------------------------------------------------------

    def generar_plantilla_vacia(self, output_path: str) -> str:
        """
        Genera una plantilla Excel vacía a partir del template.
        - Mantiene TODAS las hojas con su formato
        - Limpia filas de datos (desde fila 3 en adelante)
        - Agrega fila de ejemplo (fila 2) con datos de muestra
        - Agrega hoja NOTAS con instrucciones detalladas
        
        Args:
            output_path: Ruta donde guardar el Excel
            
        Returns:
            Ruta del archivo generado
        """
        self._validar_template()
        
        import shutil
        
        # 1. Copiar template
        shutil.copy2(self.template_path, output_path)
        
        # 2. Abrir con openpyxl
        wb = openpyxl.load_workbook(output_path, keep_vba=False)
        
        # 3. Limpiar datos y agregar ejemplos en hojas de datos
        self._preparar_hojas_plantilla(wb)
        
        # 4. Agregar hoja NOTAS
        self._agregar_hoja_notas(wb)
        
        # 5. Guardar
        wb.save(output_path)
        wb.close()
        logger.info(f"Plantilla vacía generada: {output_path}")
        
        return output_path

    def _preparar_hojas_plantilla(self, wb):
        """
        Prepara las hojas de datos en la plantilla:
        - Limpia filas de datos desde fila 3
        - Agrega fila 2 con datos de ejemplo
        - Mantiene headers y formato intactos
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Hojas que deben tener datos de ejemplo
        HOJAS_EJEMPLO = {
            "ID lotes": {
                "ejemplo": ["Ejemplo Lote 1", 1.5, 7500, "Castillo", "", "", 3, "", ""],
                "limpiar_desde": 2,
            },
            "Ingresos por ventas de cafe": {
                "ejemplo": ["", "", "CPS", 150, None, 0],
                "limpiar_desde": 3,
            },
            "Instalacion de Cafe": {
                "mo": ["Ejemplo Lote 1", None, "Trazo y ahoyado", 20, 50000, None],
                "insumos": [None, "Fertilizante 15-15-15", 10, 120000, None],
                "limpiar_desde": 3,
            },
            "Control de arvenses": {
                "mo": ["Ejemplo Lote 1", None, "Control manual de arvenses", 8, 45000, None],
                "insumos": [None, "Control químico", "Glifosato", 4, 25000, None],
                "limpiar_desde": 3,
            },
            "Fertilizacion": {
                "mo": ["Ejemplo Lote 1", None, "Aplicación de fertilizante", 10, 50000, None],
                "insumos": [None, "Urea", 100, 2800, None],
                "limpiar_desde": 3,
            },
            "Control Fitosanitario": {
                "mo": ["Ejemplo Lote 1", None, "Aplicación de fungicida", 5, 45000, None],
                "insumos": [None, "Aplicación fitosanitaria", "Fungicida cúprico", 3, 35000, None],
                "limpiar_desde": 3,
            },
            "Regulacion de sombrio": {
                "mo": ["Ejemplo Lote 1", None, "Regulación de sombra", 6, 45000, None],
                "insumos": [None, "Machete", 2, 12000, None],
                "limpiar_desde": 3,
            },
            "Otras Labores": {
                "mo": ["Ejemplo Lote 1", None, "Mantenimiento general", 5, 45000, None],
                "insumos": [None, "Herramientas varias", 1, 85000, None],
                "limpiar_desde": 3,
            },
            "Recoleccion": {
                "ejemplo": [None, "recolección cereza", 200, None, 0],
                "limpiar_desde": 3,
            },
            "Beneficio": {
                "ejemplo": [None, "Beneficio húmedo", 10, 25000, None],
                "limpiar_desde": 3,
            },
            "Gastos Administrativos": {
                "ejemplo": [None, "Pago servicios públicos", 0],
                "limpiar_desde": 3,
            },
        }
        
        for hoja_nombre, config in HOJAS_EJEMPLO.items():
            if hoja_nombre not in wb.sheetnames:
                logger.warning(f"Hoja '{hoja_nombre}' no encontrada en template, saltando")
                continue
            
            ws = wb[hoja_nombre]
            limpiar_desde = config.get("limpiar_desde", 3)
            max_col = ws.max_column
            
            # Limpiar filas de datos desde 'limpiar_desde' en adelante
            for row in range(limpiar_desde, ws.max_row + 1):
                for col in range(1, max_col + 1):
                    cell = ws.cell(row=row, column=col)
                    # Verificar si la celda está fusionada (no se puede escribir en merged cells hijas)
                    try:
                        cell.value = None
                    except AttributeError:
                        # MergedCell — ignorar
                        pass
            
            # Agregar fila de ejemplo (fila 2)
            estilo_ejemplo = Font(italic=True, color="999999", size=10, name="Calibri")
            
            def _escribir_si_posible(ws, row, col, valor, font):
                """Escribe un valor en una celda si no está fusionada."""
                try:
                    cell = ws.cell(row=row, column=col, value=valor)
                    cell.font = font
                except AttributeError:
                    # MergedCell — ignorar
                    pass
            
            if "ejemplo" in config:
                for col_idx, valor in enumerate(config["ejemplo"], 1):
                    if valor is not None:
                        _escribir_si_posible(ws, 2, col_idx, valor, estilo_ejemplo)
            
            if "mo" in config:
                for col_idx, valor in enumerate(config["mo"], 1):
                    if valor is not None:
                        _escribir_si_posible(ws, 2, col_idx, valor, estilo_ejemplo)
            
            if "insumos" in config:
                for col_idx, valor in enumerate(config["insumos"], 1):
                    col_real = 8 + col_idx - 1  # Insumos empiezan en col H (8)
                    if valor is not None and col_real <= max_col:
                        _escribir_si_posible(ws, 2, col_real, valor, estilo_ejemplo)
            
            logger.debug(f"Hoja '{hoja_nombre}' preparada en plantilla vacía")

    def _agregar_hoja_notas(self, wb):
        """
        Agrega o actualiza la hoja NOTAS con instrucciones detalladas
        para llenar la plantilla.
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # Eliminar si ya existe para recrearla
        if 'NOTAS' in wb.sheetnames:
            del wb['NOTAS']
        
        ws_notas = wb.create_sheet('NOTAS', 0)  # Primera posición
        
        notas = [
            ("📋 INSTRUCCIONES PARA LLENAR LA PLANTILLA", True, 14, "1F4E79"),
            ("", False, 10, "000000"),
            ("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", False, 10, "2E7D32"),
            ("📄 HOJA: Fincas (solo para importación — no está en este Excel)", True, 11, "000000"),
            ("   • Crea las fincas desde el bot con /fincas", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("📄 HOJA: ID lotes", True, 11, "000000"),
            ("   • Agrega aquí los lotes de tu finca", False, 10, "000000"),
            ("   • nombre: Nombre del lote (ej: Lote La Esperanza 1)", False, 10, "000000"),
            ("   • area_hectareas: Área en hectáreas (ej: 1.5)", False, 10, "000000"),
            ("   • num_arboles: Número de árboles (ej: 7500)", False, 10, "000000"),
            ("   • variedad: Castillo, Caturra, Colombia, etc.", False, 10, "000000"),
            ("   • fecha_siembra: Formato DD/MM/AAAA o AAAA-MM-DD", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("📄 HOJA: Ingresos por ventas de cafe", True, 11, "000000"),
            ("   • Tipo café: CPS (Pergamino Seco), Pasilla", False, 10, "000000"),
            ("   • Cantidad (kg): Kilos vendidos (número)", False, 10, "000000"),
            ("   • Valor Unitario: Se calcula automáticamente (Valor Total / Cantidad)", False, 10, "000000"),
            ("   • Valor Total: Valor total de la venta en COP", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("📄 HOJAS DE COSTOS (Instalacion, Arvenses, Fertilizacion, etc.)", True, 11, "000000"),
            ("   ESTRUCTURA: Col A-F = Mano de Obra (MO), Col H+ = Insumos", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("   SECCIÓN MO (Columnas A-F):", True, 10, "1565C0"),
            ("   • Lote: Nombre del lote donde se realizó la labor", False, 10, "000000"),
            ("   • Fecha: Fecha de la labor (DD/MM/AAAA)", False, 10, "000000"),
            ("   • Labor: Descripción de la labor realizada", False, 10, "000000"),
            ("   • Cantidad: Número de jornales o unidades", False, 10, "000000"),
            ("   • Valor Unitario: Costo por jornal/unidad en COP", False, 10, "000000"),
            ("   • Valor Total: Se calcula automáticamente (Cant. × V.Unitario)", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("   SECCIÓN INSUMOS (Columnas H+):", True, 10, "C62828"),
            ("   • Fecha: Fecha de la compra/aplicación", False, 10, "000000"),
            ("   • Producto: Nombre del producto/insumo", False, 10, "000000"),
            ("   • Cantidad: Cantidad adquirida", False, 10, "000000"),
            ("   • Valor Unitario: Precio por unidad en COP", False, 10, "000000"),
            ("   • Valor Total: Se calcula automáticamente (Cant. × V.Unitario)", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("📄 HOJA: Recoleccion", True, 11, "000000"),
            ("   • Fecha: Fecha de recolección", False, 10, "000000"),
            ("   • Labor: Descripción (ej: recolección cereza)", False, 10, "000000"),
            ("   • Kilos: Kilos de café cereza recolectados", False, 10, "000000"),
            ("   • V.Unitario: Se calcula automáticamente (V.Total / Kilos)", False, 10, "000000"),
            ("   • V.Total: Valor total pagado por la recolección", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("📄 HOJA: Beneficio", True, 11, "000000"),
            ("   • Fecha: Fecha del beneficio", False, 10, "000000"),
            ("   • Labor: Descripción (ej: Beneficio húmedo)", False, 10, "000000"),
            ("   • Jornales: Número de jornales", False, 10, "000000"),
            ("   • V.Unitario: Valor por jornal", False, 10, "000000"),
            ("   • V.Total: Se calcula automáticamente (Jornales × V.Unitario)", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("📄 HOJA: Gastos Administrativos", True, 11, "000000"),
            ("   • Fecha: Fecha del gasto", False, 10, "000000"),
            ("   • Gasto: Descripción del gasto administrativo", False, 10, "000000"),
            ("   • V.Total: Valor total del gasto", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", False, 10, "C62828"),
            ("⚠️ IMPORTANTE:", True, 12, "C62828"),
            ("   • No modifiques los nombres de las columnas (fila 1)", False, 10, "000000"),
            ("   • No agregues ni elimines hojas", False, 10, "000000"),
            ("   • Los valores numéricos usan punto como decimal (ej: 1.5)", False, 10, "000000"),
            ("   • Usa el formato de fecha DD/MM/AAAA o AAAA-MM-DD", False, 10, "000000"),
            ("   • Las fórmulas en Valor Total se calculan automáticamente", False, 10, "000000"),
            ("   • Guarda el archivo y súbelo al bot con /importar", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━", False, 10, "2E7D32"),
            ("💡 EJEMPLO RÁPIDO:", True, 12, "2E7D32"),
            ("   1. Crea una finca con /fincas", False, 10, "000000"),
            ("   2. Abre este Excel (la fila 2 tiene datos de ejemplo como guía)", False, 10, "000000"),
            ("   3. Reemplaza los datos de ejemplo con tus datos reales", False, 10, "000000"),
            ("   4. Guarda el archivo", False, 10, "000000"),
            ("   5. Subilo al bot con /importar", False, 10, "000000"),
            ("", False, 10, "000000"),
            ("📌 Los datos se importarán a tu cuenta automáticamente.", False, 10, "000000"),
        ]
        
        for i, (texto, negrita, tamano, color) in enumerate(notas, 1):
            cell = ws_notas.cell(row=i, column=1, value=texto)
            cell.font = Font(bold=negrita, size=tamano, color=color, name="Calibri")
            cell.alignment = Alignment(vertical="center")
        
        ws_notas.column_dimensions['A'].width = 85
        ws_notas.sheet_properties.tabColor = "1F4E79"
        
        logger.info("Hoja NOTAS agregada a la plantilla vacía")

    # ------------------------------------------------------------------
    # Método principal
    # ------------------------------------------------------------------

    def generar_excel(self, finca_id: int, db, output_path: str) -> str:
        """
        Genera el Excel con datos de la finca.
        1. Copia el template a output_path
        2. Abre con openpyxl
        3. Llena SOLO hojas de datos
        4. Guarda (manteniendo fórmulas)
        
        Args:
            finca_id: ID de la finca
            db: Instancia de Database
            output_path: Ruta donde guardar el Excel
            
        Returns:
            Ruta del archivo generado
        """
        self._validar_template()

        # Obtener todos los datos de la finca
        data = db.get_all_data_for_export(finca_id)
        finca = db.get_finca_by_id(finca_id)

        logger.info(f"Generando Excel para finca '{finca['nombre']}' (ID: {finca_id})")

        # 1. Copiar template
        shutil.copy2(self.template_path, output_path)
        logger.info(f"Template copiado a: {output_path}")

        # 2. Abrir con openpyxl (keep_vba=False para limpiar macros si hubiera)
        wb = openpyxl.load_workbook(output_path, keep_vba=False)
        logger.info(f"Hojas disponibles: {wb.sheetnames}")

        # 3. Llenar hojas
        self._llenar_hoja_lotes(wb, data.get("lotes", []))
        self._llenar_hoja_ingresos(wb, data)
        self._llenar_hojas_costos(wb, data)

        # 3b. Llenar hoja Presupuesto (si existe en template o crearla)
        self._llenar_hoja_presupuesto(wb, db, finca_id)

        # 3c. Llenar hoja Indicadores
        self._llenar_hoja_indicadores(wb, db, finca_id)

        # 4. Generar hoja de gráficos
        ws_graficos = wb.create_sheet("Gráficos")
        # Mover "Gráficos" después de "ID lotes"
        idx_lotes = wb.sheetnames.index("ID lotes") if "ID lotes" in wb.sheetnames else 0
        wb.move_sheet("Gráficos", offset=idx_lotes + 1 - len(wb.sheetnames) + 1)
        self._generar_hoja_graficos(db, finca_id, ws_graficos)

        # 5. Guardar
        wb.save(output_path)
        wb.close()
        logger.info(f"Excel generado exitosamente: {output_path}")

        return output_path

    # ------------------------------------------------------------------
    # Hoja ID lotes
    # ------------------------------------------------------------------

    def _llenar_hoja_lotes(self, wb, lotes: list):
        """Llena la hoja 'ID lotes' con los datos de los lotes."""
        if "ID lotes" not in wb.sheetnames:
            logger.warning("Hoja 'ID lotes' no encontrada en el template")
            return

        ws = wb["ID lotes"]
        logger.info(f"Llenando hoja 'ID lotes' con {len(lotes)} lotes")

        num_needed = len(lotes)
        DATA_START = HOJA_LOTES_START_ROW  # 2
        SUBTOTAL_ROW = HOJA_LOTES_SUBTOTAL_ROW  # 17
        TEMPLATE_ROWS = HOJA_LOTES_TEMPLATE_ROWS  # 15
        MAX_COL = 9  # Columnas en esta hoja (A-I)

        # Asegurar filas suficientes
        new_subtotal_row, last_data_row = self._asegurar_filas_suficientes(
            ws, DATA_START, SUBTOTAL_ROW, num_needed, TEMPLATE_ROWS, MAX_COL
        )

        # Llenar datos
        for i, lote in enumerate(lotes):
            fila = DATA_START + i
            if fila >= new_subtotal_row:  # Safety check
                logger.warning(f"No hay espacio para lote: {lote['nombre']}")
                break

            # Fórmula de EDAD (col 6) — copiar si no existe
            col_f = ws.cell(row=fila, column=6)
            if not col_f.value or not str(col_f.value).startswith("="):
                col_f.value = f'=IF(E{fila}="", 0, DATEDIF(VALUE(E{fila}), TODAY(), "M"))'

            ws.cell(row=fila, column=1, value=lote["nombre"])        # A: ID LOTE
            ws.cell(row=fila, column=2, value=lote["area_hectareas"] or 0)  # B: AREA

            arboles = lote.get("num_arboles") or 0
            try:
                ws.cell(row=fila, column=3, value=int(arboles))      # C: # ARBOLES
            except (ValueError, TypeError):
                ws.cell(row=fila, column=3, value=0)

            ws.cell(row=fila, column=4, value=lote["variedad"] or "")  # D: VARIEDAD

            # E: Fecha de siembra
            fecha = lote.get("fecha_siembra", "") or ""
            if fecha:
                try:
                    from datetime import datetime as dt
                    fecha_dt = dt.strptime(fecha, "%Y-%m-%d")
                    ws.cell(row=fila, column=5, value=fecha_dt)
                    ws.cell(row=fila, column=5).number_format = "DD/MM/YYYY"
                except ValueError:
                    ws.cell(row=fila, column=5, value=fecha)
            else:
                ws.cell(row=fila, column=5, value="")

        logger.info(f"Hoja 'ID lotes' llenada: {num_needed} lotes (subtotal fila {new_subtotal_row})")

    # ------------------------------------------------------------------
    # Hoja Ingresos
    # ------------------------------------------------------------------

    def _llenar_hoja_ingresos(self, wb, data: dict):
        """Llena la hoja 'Ingresos por ventas de cafe'."""
        hoja_nombre = "Ingresos por ventas de cafe"
        if hoja_nombre not in wb.sheetnames:
            logger.warning(f"Hoja '{hoja_nombre}' no encontrada")
            return

        ws = wb[hoja_nombre]

        # Recolectar todos los ingresos
        ingresos = []
        for cat in ["ingreso_cps", "ingreso_pasilla"]:
            for t in data.get(cat, []):
                tipo_map = {"ingreso_cps": "CPS", "ingreso_pasilla": "Pasilla"}
                ingresos.append({
                    "fecha": t["fecha"],
                    "tipo": tipo_map[cat],
                    "cantidad": t["cantidad"] or 0,
                    "valor_total": t["valor_total"] or 0,
                })

        logger.info(f"Llenando hoja '{hoja_nombre}' con {len(ingresos)} ingresos")

        num_needed = len(ingresos)
        DATA_START = HOJA_COSTOS_START_ROW  # 3
        SUBTOTAL_ROW = HOJA_COSTOS_SUBTOTAL_ROW  # 20
        TEMPLATE_ROWS = HOJA_COSTOS_TEMPLATE_ROWS  # 17
        MAX_COL = 14  # Columnas en esta hoja (A-N)

        # Asegurar filas suficientes
        new_subtotal_row, last_data_row = self._asegurar_filas_suficientes(
            ws, DATA_START, SUBTOTAL_ROW, num_needed, TEMPLATE_ROWS, MAX_COL
        )

        # Actualizar fórmulas SUM en subtotal (columnas B, D, F)
        self._actualizar_sum_subtotal(
            ws, new_subtotal_row, DATA_START, last_data_row, celdas_sum=[2, 4, 6]
        )

        # Llenar datos
        for i, ingreso in enumerate(ingresos):
            fila = DATA_START + i
            if fila >= new_subtotal_row:
                break

            # Extender fórmula de valor unitario si no existe
            col_e = ws.cell(row=fila, column=5)
            if not col_e.value or not str(col_e.value).startswith("="):
                col_e.value = f'=IFERROR(F{fila}/D{fila}, 0)'

            # A: Fecha
            fecha = ingreso.get("fecha", "")
            if fecha:
                try:
                    from datetime import datetime as dt
                    fecha_dt = dt.strptime(fecha, "%Y-%m-%d")
                    ws.cell(row=fila, column=1, value=fecha_dt)
                    ws.cell(row=fila, column=1).number_format = "DD/MM/YYYY"
                except ValueError:
                    ws.cell(row=fila, column=1, value=fecha)
            else:
                ws.cell(row=fila, column=1, value="")

            # B: Unidad de venta (dejar vacío)
            ws.cell(row=fila, column=2, value="")

            # C: Tipo de café
            ws.cell(row=fila, column=3, value=ingreso["tipo"])

            # D: Cantidad
            ws.cell(row=fila, column=4, value=ingreso["cantidad"])

            # E: Valor Unitario - DEJAR LA FÓRMULA (=F/D)

            # F: Valor total
            ws.cell(row=fila, column=6, value=ingreso["valor_total"])

        logger.info(f"Hoja '{hoja_nombre}' llenada con {num_needed} registros")

    # ------------------------------------------------------------------
    # Hojas de costos (MO + Insumos, y especiales)
    # ------------------------------------------------------------------

    def _llenar_hojas_costos(self, wb, data: dict):
        """Llena todas las hojas de costos."""
        # Hojas con estructura MO + Insumos
        hojas_mo_insumos = [
            "Instalacion de Cafe",
            "Control de arvenses",
            "Fertilizacion",
            "Control Fitosanitario",
            "Regulacion de sombrio",
            "Otras Labores",
        ]

        for hoja in hojas_mo_insumos:
            if hoja not in wb.sheetnames:
                logger.warning(f"Hoja '{hoja}' no encontrada, saltando")
                continue
            self._llenar_hoja_mo_insumos(wb, hoja, data)

        # Hojas especiales
        if "Recoleccion" in wb.sheetnames:
            self._llenar_hoja_recoleccion(wb, data)
        if "Beneficio" in wb.sheetnames:
            self._llenar_hoja_beneficio(wb, data)
        if "Gastos Administrativos" in wb.sheetnames:
            self._llenar_hoja_gastos_admin(wb, data)

    def _llenar_hoja_mo_insumos(self, wb, hoja_nombre: str, data: dict):
        """
        Llena una hoja que tiene estructura MO (A-F) e Insumos (H-M).
        
        Estrategia: MO e insumos se llenan en las mismas filas, desde la fila 3.
        Si hay más MO que insumos (o viceversa), las celdas vacías quedan en blanco.
        """
        config = HOJA_CONFIG.get(hoja_nombre)
        if not config:
            logger.warning(f"Sin configuración para hoja '{hoja_nombre}'")
            return

        ws = wb[hoja_nombre]

        # Recolectar datos MO
        mo_data = []
        for cat in config.get("categorias_mo", []):
            mo_data.extend(data.get(cat, []))

        # Recolectar datos Insumos
        insumos_data = []
        for cat in config.get("categorias_insumos", []):
            insumos_data.extend(data.get(cat, []))

        logger.info(
            f"Llenando hoja '{hoja_nombre}': {len(mo_data)} MO, {len(insumos_data)} Insumos"
        )

        len_max = max(len(mo_data), len(insumos_data), 0)
        if len_max == 0:
            return

        mo_cols = config.get("mo_cols", {})
        insumos_cols = config.get("insumos_cols", {})

        num_needed = len_max
        DATA_START = HOJA_COSTOS_START_ROW  # 3
        SUBTOTAL_ROW = HOJA_COSTOS_SUBTOTAL_ROW  # 20
        TEMPLATE_ROWS = HOJA_COSTOS_TEMPLATE_ROWS  # 17
        MAX_COL = 20  # Suficiente para todas las variantes (Control arvenses usa hasta col T)

        # Asegurar filas suficientes
        new_subtotal_row, last_data_row = self._asegurar_filas_suficientes(
            ws, DATA_START, SUBTOTAL_ROW, num_needed, TEMPLATE_ROWS, MAX_COL
        )

        # Determinar qué columnas SUM están en subtotal según configuración de la hoja
        # MO subtotal: col 6 (suma F)
        # Insumos subtotal: depende de la hoja
        celdas_sum_mo = [6]  # Siempre col F
        celdas_sum_insumos = []
        if insumos_cols:
            # La columna de valor total de insumos está al final del rango de insumos
            ins_start_letter = insumos_cols["start"]
            ins_end_letter = insumos_cols["end"]
            ins_total_col = ord(ins_end_letter) - ord("A") + 1
            celdas_sum_insumos = [ins_total_col]

        # Actualizar fórmulas SUM en subtotal para MO e Insumos
        self._actualizar_sum_subtotal(
            ws, new_subtotal_row, DATA_START, last_data_row,
            celdas_sum=celdas_sum_mo + celdas_sum_insumos
        )

        # Llenar datos
        for i in range(len_max):
            fila = DATA_START + i
            if fila >= new_subtotal_row:
                logger.warning(f"Se alcanzó el límite de filas en '{hoja_nombre}'")
                break

            # --- Llenar MO (columnas A-F) ---
            if i < len(mo_data):
                self._escribir_fila_mo(ws, fila, mo_data[i], mo_cols)
            else:
                # Asegurar fórmula de V.Total MO incluso si no hay datos MO en esta fila
                self._asegurar_formula_mo(ws, fila, mo_cols)

            # --- Llenar Insumos (columnas H-M) ---
            if i < len(insumos_data):
                self._escribir_fila_insumos(ws, fila, insumos_data[i], insumos_cols)
            else:
                # Asegurar fórmula de V.Total Insumos incluso si no hay datos
                self._asegurar_formula_insumos(ws, fila, insumos_cols)

        logger.info(
            f"Hoja '{hoja_nombre}' llenada: {len(mo_data)} MO, {len(insumos_data)} Insumos"
        )

    def _asegurar_formula_mo(self, ws, fila: int, cols_config: dict):
        """Asegura que exista la fórmula de V.Total en la columna F de una fila MO."""
        if not cols_config:
            return
        end_col = ord(cols_config["end"]) - ord("A") + 1
        celda = ws.cell(row=fila, column=end_col)
        if not celda.value or not str(celda.value).startswith("="):
            # Fórmula típica: =IFERROR(D{fila}*E{fila}, 0) o =+E{fila}*D{fila}
            # Determinar según la configuración
            campos = cols_config.get("campos", [])
            # Encontrar columnas de cantidad y valor_unitario
            col_cant = None
            col_vu = None
            for j, campo in enumerate(campos):
                start_col = ord(cols_config["start"]) - ord("A") + 1
                if campo == "cantidad":
                    col_cant = start_col + j
                elif campo == "valor_unitario":
                    col_vu = start_col + j
            if col_cant and col_vu:
                celda.value = f'=IFERROR({get_column_letter(col_vu)}{fila}*{get_column_letter(col_cant)}{fila}, 0)'

    def _asegurar_formula_insumos(self, ws, fila: int, cols_config: dict):
        """Asegura que exista la fórmula de V.Total en la última columna de insumos."""
        if not cols_config:
            return
        end_col = ord(cols_config["end"]) - ord("A") + 1
        celda = ws.cell(row=fila, column=end_col)
        if not celda.value or not str(celda.value).startswith("="):
            # Fórmula típica: =IFERROR(+K{fila}*J{fila}, 0)
            campos = cols_config.get("campos", [])
            col_cant = None
            col_vu = None
            for j, campo in enumerate(campos):
                start_col = ord(cols_config["start"]) - ord("A") + 1
                if campo == "cantidad":
                    col_cant = start_col + j
                elif campo == "valor_unitario":
                    col_vu = start_col + j
            if col_cant and col_vu:
                celda.value = f'=IFERROR(+{get_column_letter(col_cant)}{fila}*{get_column_letter(col_vu)}{fila}, 0)'

    def _escribir_fila_mo(self, ws, fila: int, record: dict, cols_config: dict):
        """Escribe un registro de MO en una fila."""
        if not cols_config:
            return
        start_col = ord(cols_config["start"]) - ord("A") + 1  # 1-indexed
        campos = cols_config["campos"]

        for j, campo in enumerate(campos):
            col = start_col + j
            if campo == "lote":
                lote_id = record.get("lote_id", 0)
                ws.cell(row=fila, column=col, value=str(lote_id) if lote_id else "")
            elif campo == "fecha":
                self._poner_fecha(ws, fila, col, record.get("fecha", ""))
            elif campo == "labor":
                ws.cell(row=fila, column=col, value=record.get("labor", ""))
            elif campo == "cantidad":
                ws.cell(row=fila, column=col, value=record.get("cantidad", 0) or 0)
            elif campo == "valor_unitario":
                ws.cell(row=fila, column=col, value=record.get("valor_unitario", 0) or 0)
            elif campo == "valor_total_formula":
                # ✅ NUNCA sobreescribir la fórmula del template (F = D×E)
                pass
            else:
                ws.cell(row=fila, column=col, value="")

    def _escribir_fila_insumos(self, ws, fila: int, record: dict, cols_config: dict):
        """Escribe un registro de Insumos en una fila."""
        if not cols_config:
            return
        start_col = ord(cols_config["start"]) - ord("A") + 1  # 1-indexed
        campos = cols_config["campos"]

        for j, campo in enumerate(campos):
            col = start_col + j
            if campo == "fecha":
                self._poner_fecha(ws, fila, col, record.get("fecha", ""))
            elif campo == "labor":
                ws.cell(row=fila, column=col, value=record.get("labor", ""))
            elif campo == "producto":
                ws.cell(row=fila, column=col, value=record.get("producto", ""))
            elif campo == "cantidad":
                ws.cell(row=fila, column=col, value=record.get("cantidad", 0) or 0)
            elif campo == "valor_unitario":
                ws.cell(row=fila, column=col, value=record.get("valor_unitario", 0) or 0)
            elif campo == "valor_total_formula":
                # ✅ NUNCA sobreescribir la fórmula del template (L/M = K×J)
                pass
            else:
                ws.cell(row=fila, column=col, value="")

    # ------------------------------------------------------------------
    # Hojas especiales
    # ------------------------------------------------------------------

    def _llenar_hoja_recoleccion(self, wb, data: dict):
        """Llena la hoja 'Recoleccion'."""
        ws = wb["Recoleccion"]
        records = data.get("recoleccion", [])
        logger.info(f"Llenando hoja 'Recoleccion' con {len(records)} registros")

        num_needed = len(records)
        DATA_START = HOJA_COSTOS_START_ROW  # 3
        SUBTOTAL_ROW = HOJA_COSTOS_SUBTOTAL_ROW  # 20
        TEMPLATE_ROWS = HOJA_COSTOS_TEMPLATE_ROWS  # 17
        MAX_COL = 14  # A-N

        # Asegurar filas suficientes
        new_subtotal_row, last_data_row = self._asegurar_filas_suficientes(
            ws, DATA_START, SUBTOTAL_ROW, num_needed, TEMPLATE_ROWS, MAX_COL
        )

        # Actualizar SUBTOTAL: col C (SUM kilos), col E (SUM valor total)
        self._actualizar_sum_subtotal(
            ws, new_subtotal_row, DATA_START, last_data_row, celdas_sum=[3, 5]
        )

        # Llenar datos
        for i, rec in enumerate(records):
            fila = DATA_START + i
            if fila >= new_subtotal_row:
                break

            # A: Fecha
            self._poner_fecha(ws, fila, 1, rec.get("fecha", ""))

            # B: Recolección cereza (labor)
            ws.cell(row=fila, column=2, value=rec.get("labor", "recolección cereza"))

            # C: Kilos (cantidad)
            ws.cell(row=fila, column=3, value=rec.get("cantidad", 0) or 0)

            # D: V.Unitario - asegurar fórmula =E/C
            col_d = ws.cell(row=fila, column=4)
            if not col_d.value or not str(col_d.value).startswith("="):
                col_d.value = f'=IFERROR(E{fila}/C{fila}, 0)'

            # E: V.Total (valor_total)
            ws.cell(row=fila, column=5, value=rec.get("valor_total", 0) or 0)

        logger.info(f"Hoja 'Recoleccion' llenada con {num_needed} registros")

    def _llenar_hoja_beneficio(self, wb, data: dict):
        """Llena la hoja 'Beneficio'."""
        ws = wb["Beneficio"]
        records = data.get("beneficio", [])
        logger.info(f"Llenando hoja 'Beneficio' con {len(records)} registros")

        num_needed = len(records)
        DATA_START = HOJA_COSTOS_START_ROW  # 3
        SUBTOTAL_ROW = HOJA_COSTOS_SUBTOTAL_ROW  # 20
        TEMPLATE_ROWS = HOJA_COSTOS_TEMPLATE_ROWS  # 17
        MAX_COL = 6  # A-F

        # Asegurar filas suficientes
        new_subtotal_row, last_data_row = self._asegurar_filas_suficientes(
            ws, DATA_START, SUBTOTAL_ROW, num_needed, TEMPLATE_ROWS, MAX_COL
        )

        # Actualizar SUBTOTAL: col E (SUM valor total)
        self._actualizar_sum_subtotal(
            ws, new_subtotal_row, DATA_START, last_data_row, celdas_sum=[5]
        )

        # Llenar datos
        for i, rec in enumerate(records):
            fila = DATA_START + i
            if fila >= new_subtotal_row:
                break

            # A: Fecha
            self._poner_fecha(ws, fila, 1, rec.get("fecha", ""))

            # B: Labor realizada
            ws.cell(row=fila, column=2, value=rec.get("labor", ""))

            # C: Número de jornales (cantidad)
            ws.cell(row=fila, column=3, value=rec.get("cantidad", 0) or 0)

            # D: V.Unitario (valor_unitario)
            ws.cell(row=fila, column=4, value=rec.get("valor_unitario", 0) or 0)

            # E: V.Total - asegurar fórmula =D*C
            col_e = ws.cell(row=fila, column=5)
            if not col_e.value or not str(col_e.value).startswith("="):
                col_e.value = f'=IFERROR(D{fila}*C{fila}, 0)'

        logger.info(f"Hoja 'Beneficio' llenada con {num_needed} registros")

    def _llenar_hoja_gastos_admin(self, wb, data: dict):
        """Llena la hoja 'Gastos Administrativos'."""
        ws = wb["Gastos Administrativos"]
        records = data.get("administrativo", [])
        logger.info(f"Llenando hoja 'Gastos Administrativos' con {len(records)} registros")

        num_needed = len(records)
        DATA_START = HOJA_COSTOS_START_ROW  # 3
        SUBTOTAL_ROW = HOJA_COSTOS_SUBTOTAL_ROW  # 20
        TEMPLATE_ROWS = HOJA_COSTOS_TEMPLATE_ROWS  # 17
        MAX_COL = 9  # A-I

        # Asegurar filas suficientes
        new_subtotal_row, last_data_row = self._asegurar_filas_suficientes(
            ws, DATA_START, SUBTOTAL_ROW, num_needed, TEMPLATE_ROWS, MAX_COL
        )

        # Actualizar SUBTOTAL: col C (SUM valor total)
        self._actualizar_sum_subtotal(
            ws, new_subtotal_row, DATA_START, last_data_row, celdas_sum=[3]
        )

        # Llenar datos
        for i, rec in enumerate(records):
            fila = DATA_START + i
            if fila >= new_subtotal_row:
                break

            # A: Fecha
            self._poner_fecha(ws, fila, 1, rec.get("fecha", ""))

            # B: Gasto administrativo (labor)
            ws.cell(row=fila, column=2, value=rec.get("labor", ""))

            # C: V.Total (valor_total)
            ws.cell(row=fila, column=3, value=rec.get("valor_total", 0) or 0)

        logger.info(f"Hoja 'Gastos Administrativos' llenada con {num_needed} registros")

    # ------------------------------------------------------------------
    # Hoja Presupuesto
    # ------------------------------------------------------------------

    def _llenar_hoja_presupuesto(self, wb, db, finca_id: int):
        """Llena la hoja 'Presupuesto' con planificación vs ejecución.
        
        Estructura:
        A: Categoría | B: % Referencia del sector | C: Monto Planificado | D: Monto Ejecutado 
        | E: Diferencia | F: % Ejecución
        
        Usa colores: Verde si ejecutado <= planificado, Rojo si sobregiro.
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Definir colores
        FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        FILL_TOTAL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        FILL_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        FILL_ROJO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        FONT_HEADER = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        FONT_TITLE = Font(bold=True, size=14, name="Calibri", color="1F4E79")
        FONT_NORMAL = Font(size=10, name="Calibri")
        FONT_BOLD = Font(bold=True, size=10, name="Calibri")
        THIN_BORDER = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        CENTER = Alignment(horizontal="center", vertical="center")
        RIGHT = Alignment(horizontal="right", vertical="center")
        LEFT = Alignment(horizontal="left", vertical="center")

        # Estructura de costos del sector
        FEPCAFE = [
            ("recoleccion", "☕ Recolección", 54),
            ("fertilizacion", "🧪 Fertilización", 19),
            ("administrativo", "📋 Gastos Admin/Financieros", 7),
            ("arvenses", "🌿 Manejo de Arvenses", 6),
            ("beneficio", "🏭 Beneficio", 6),
            ("instalacion", "🌱 Renovación", 5),
            ("fitosanitario", "🛡️ Fitosanitarios", 2),
            ("otras_labores", "🔧 Otras labores", 1),
        ]

        # Determinar año más reciente con presupuesto
        anios = db.get_presupuesto_anios(finca_id)
        if not anios:
            logger.info("No hay presupuestos guardados, saltando hoja Presupuesto")
            return

        anio = anios[0]  # Más reciente
        presupuesto_data = db.get_ejecucion_presupuesto(finca_id, anio)
        categorias_data = {c["categoria"]: c for c in presupuesto_data["categorias"]}

        # Crear o reemplazar la hoja
        hoja_nombre = "Presupuesto"
        if hoja_nombre in wb.sheetnames:
            del wb[hoja_nombre]
        ws = wb.create_sheet(hoja_nombre)

        # ─── Título ───
        ws.merge_cells("A1:F1")
        cell = ws.cell(row=1, column=1, value=f"Presupuesto {anio}")
        cell.font = FONT_TITLE
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

        # ─── Encabezados (fila 3) ───
        headers = ["Categoría", "% Referencia", "Monto Planificado", "Monto Ejecutado", "Diferencia", "% Ejecución"]
        col_widths = [30, 12, 20, 20, 20, 15]
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            letter = chr(64 + col_idx)
            ws.column_dimensions[letter].width = width

        # ─── Datos ───
        DATA_START_ROW = 4
        row = DATA_START_ROW
        for cat_id, nombre_cat, pct_ref in FEPCAFE:
            cat_data = categorias_data.get(cat_id, {"monto_planificado": 0, "monto_ejecutado": 0, "diferencia": 0, "pct_ejecucion": 0})
            plan = cat_data["monto_planificado"]
            ejec = cat_data["monto_ejecutado"]
            diff = cat_data["diferencia"]
            pct_ejec = cat_data["pct_ejecucion"]

            # Determinar color
            fill_row = FILL_VERDE if ejec <= plan else FILL_ROJO

            ws.cell(row=row, column=1, value=nombre_cat).font = FONT_NORMAL
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=1).border = THIN_BORDER

            ws.cell(row=row, column=2, value=pct_ref).font = FONT_NORMAL
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=2).number_format = '0"%"'

            ws.cell(row=row, column=3, value=plan).font = FONT_NORMAL
            ws.cell(row=row, column=3).alignment = RIGHT
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.cell(row=row, column=3).number_format = '$#,##0'

            ws.cell(row=row, column=4, value=ejec).font = FONT_NORMAL
            ws.cell(row=row, column=4).alignment = RIGHT
            ws.cell(row=row, column=4).border = THIN_BORDER
            ws.cell(row=row, column=4).number_format = '$#,##0'

            ws.cell(row=row, column=5, value=diff).font = FONT_NORMAL
            ws.cell(row=row, column=5).alignment = RIGHT
            ws.cell(row=row, column=5).border = THIN_BORDER
            ws.cell(row=row, column=5).number_format = '$#,##0'
            ws.cell(row=row, column=5).fill = fill_row

            ws.cell(row=row, column=6, value=pct_ejec / 100).font = FONT_NORMAL
            ws.cell(row=row, column=6).alignment = CENTER
            ws.cell(row=row, column=6).border = THIN_BORDER
            ws.cell(row=row, column=6).number_format = '0.0%'
            ws.cell(row=row, column=6).fill = fill_row

            # Aplicar color también a cols 3 y 4 si hay sobregiro
            if ejec > plan:
                ws.cell(row=row, column=3).fill = FILL_ROJO
                ws.cell(row=row, column=4).fill = FILL_ROJO

            row += 1

        # ─── Fila Total ───
        pct_total = (presupuesto_data["total_ejecutado"] / presupuesto_data["total_planificado"] * 100) if presupuesto_data["total_planificado"] > 0 else 0
        fill_total = FILL_VERDE if presupuesto_data["total_ejecutado"] <= presupuesto_data["total_planificado"] else FILL_ROJO

        ws.cell(row=row, column=1, value="TOTAL").font = FONT_BOLD
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.cell(row=row, column=1).fill = FILL_TOTAL

        ws.cell(row=row, column=2, value=100).font = FONT_BOLD
        ws.cell(row=row, column=2).alignment = CENTER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = FILL_TOTAL
        ws.cell(row=row, column=2).number_format = '0"%"'

        ws.cell(row=row, column=3, value=presupuesto_data["total_planificado"]).font = FONT_BOLD
        ws.cell(row=row, column=3).alignment = RIGHT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).fill = fill_total
        ws.cell(row=row, column=3).number_format = '$#,##0'

        ws.cell(row=row, column=4, value=presupuesto_data["total_ejecutado"]).font = FONT_BOLD
        ws.cell(row=row, column=4).alignment = RIGHT
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).fill = fill_total
        ws.cell(row=row, column=4).number_format = '$#,##0'

        ws.cell(row=row, column=5, value=presupuesto_data["total_diferencia"]).font = FONT_BOLD
        ws.cell(row=row, column=5).alignment = RIGHT
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=5).fill = fill_total
        ws.cell(row=row, column=5).number_format = '$#,##0'

        ws.cell(row=row, column=6, value=pct_total / 100).font = FONT_BOLD
        ws.cell(row=row, column=6).alignment = CENTER
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=6).fill = fill_total
        ws.cell(row=row, column=6).number_format = '0.0%'

        # ─── Gráfico de barras ───
        try:
            chart_data_row = row + 3
            ws.cell(row=chart_data_row, column=1, value="Categoría")
            ws.cell(row=chart_data_row, column=2, value="Planificado")
            ws.cell(row=chart_data_row, column=3, value="Ejecutado")

            for i, (cat_id, nombre_cat, _) in enumerate(FEPCAFE):
                r = chart_data_row + 1 + i
                cat_data = categorias_data.get(cat_id, {})
                # Usar nombre corto para el gráfico
                label = nombre_cat.split(" ")[-1] if " " in nombre_cat else nombre_cat
                ws.cell(row=r, column=1, value=label)
                ws.cell(row=r, column=2, value=cat_data.get("monto_planificado", 0))
                ws.cell(row=r, column=3, value=cat_data.get("monto_ejecutado", 0))

            chart = BarChart()
            chart.type = "col"
            chart.title = f"Planificado vs Ejecutado — {anio}"
            chart.y_axis.title = "Monto ($)"
            chart.style = 10
            chart.width = 24
            chart.height = 14

            data_ref = Reference(ws, min_col=1, min_row=chart_data_row, max_col=3, max_row=chart_data_row + len(FEPCAFE))
            cats_ref = Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=chart_data_row + len(FEPCAFE))
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)

            chart.series[0].graphicalProperties.solidFill = "1565C0"
            chart.series[1].graphicalProperties.solidFill = "C62828"

            ws.add_chart(chart, f"A{chart_data_row + len(FEPCAFE) + 2}")

        except Exception as e:
            logger.warning(f"No se pudo generar el gráfico de presupuesto: {e}")

        logger.info(f"Hoja 'Presupuesto' generada para finca {finca_id} año {anio}")

    # ------------------------------------------------------------------
    # Hoja Indicadores Técnicos
    # ------------------------------------------------------------------

    def _llenar_hoja_indicadores(self, wb, db, finca_id: int):
        """Crea la hoja 'Indicadores' con los indicadores técnicos de rendimiento.
        
        Incluye:
        - Tabla de indicadores agrupados (Área, MO, Insumos, Financiero, Productividad)
        - Gráfico de barras: MO vs Insumos por ha
        - Gráfico de barras: Productividad vs Rendimiento
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

        indicadores = db.get_indicadores_tecnicos(finca_id)
        finca = db.get_finca_by_id(finca_id)

        if not finca:
            logger.warning(f"Finca {finca_id} no encontrada para indicadores")
            return

        # Crear o reemplazar la hoja
        hoja_nombre = "Indicadores"
        if hoja_nombre in wb.sheetnames:
            del wb[hoja_nombre]
        ws = wb.create_sheet(hoja_nombre)

        # ─── Estilos ───
        FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        FILL_SECTION = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
        FILL_VERDE = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        FONT_HEADER = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        FONT_TITLE = Font(bold=True, size=14, name="Calibri", color="1F4E79")
        FONT_SECTION = Font(bold=True, size=11, name="Calibri", color="1F4E79")
        FONT_NORMAL = Font(size=10, name="Calibri")
        FONT_BOLD = Font(bold=True, size=10, name="Calibri")
        FONT_VALUE = Font(bold=True, size=11, name="Calibri", color="2E7D32")
        THIN_BORDER = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        CENTER = Alignment(horizontal="center", vertical="center")
        RIGHT = Alignment(horizontal="right", vertical="center")
        LEFT = Alignment(horizontal="left", vertical="center")

        # ─── Título ───
        ws.merge_cells("A1:D1")
        cell = ws.cell(row=1, column=1, value=f"📊 Indicadores Técnicos — {finca['nombre']}")
        cell.font = FONT_TITLE
        cell.alignment = Alignment(horizontal="center", vertical="center")

        # ─── Encabezados (fila 3) ───
        headers = ["Indicador", "Valor", "Unidad", "Referencia del sector"]
        col_widths = [35, 25, 15, 20]
        for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            letter = chr(64 + col_idx)
            ws.column_dimensions[letter].width = width

        # ─── Datos ───
        row = 4

        def escribir_indicador(ws, row, label, valor, unidad, ref_fnc=""):
            ws.cell(row=row, column=1, value=label).font = FONT_NORMAL
            ws.cell(row=row, column=1).alignment = LEFT
            ws.cell(row=row, column=1).border = THIN_BORDER

            cell_val = ws.cell(row=row, column=2, value=valor)
            cell_val.font = FONT_VALUE
            cell_val.alignment = RIGHT
            cell_val.border = THIN_BORDER
            cell_val.number_format = '#,##0.00'

            ws.cell(row=row, column=3, value=unidad).font = FONT_NORMAL
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=3).border = THIN_BORDER

            ws.cell(row=row, column=4, value=ref_fnc).font = FONT_NORMAL
            ws.cell(row=row, column=4).alignment = CENTER
            ws.cell(row=row, column=4).border = THIN_BORDER

        def escribir_seccion(ws, row, titulo):
            ws.merge_cells(f"A{row}:D{row}")
            cell = ws.cell(row=row, column=1, value=titulo)
            cell.font = FONT_SECTION
            cell.fill = FILL_SECTION
            cell.alignment = LEFT
            cell.border = THIN_BORDER
            for c in range(2, 5):
                ws.cell(row=row, column=c).fill = FILL_SECTION
                ws.cell(row=row, column=c).border = THIN_BORDER
            return row + 1

        def formatear_moneda_indicador(valor):
            """Formato moneda para el Excel (número con formato $)."""
            return round(valor, 2)

        # Sección: Área
        row = escribir_seccion(ws, row, "🌱 Área")
        escribir_indicador(ws, row, "Área Total", round(indicadores['area_total'], 2), "ha", "")
        row += 1
        escribir_indicador(ws, row, "Área Productiva", round(indicadores['area_productiva'], 2), "ha", "")
        row += 2

        # Sección: Mano de Obra
        row = escribir_seccion(ws, row, "👷 Mano de Obra")
        escribir_indicador(ws, row, "Total Jornales", round(indicadores['total_jornales'], 0), "jornales", "")
        row += 1
        escribir_indicador(ws, row, "Jornales por Hectárea", round(indicadores['jornales_por_ha'], 2), "jornales/ha", "80-120")
        row += 1
        escribir_indicador(ws, row, "Costo MO por Hectárea", formatear_moneda_indicador(indicadores['costo_mo_por_ha']), "$/ha", "")
        row += 1
        escribir_indicador(ws, row, "Eficiencia de MO", round(indicadores['eficiencia_mo'], 2), "kg/jornal", "0.3-0.5")
        row += 2

        # Sección: Insumos
        row = escribir_seccion(ws, row, "🧪 Insumos")
        escribir_indicador(ws, row, "Costo Insumos por Hectárea", formatear_moneda_indicador(indicadores['costo_insumos_por_ha']), "$/ha", "")
        row += 1
        escribir_indicador(ws, row, "Kg Producidos (CPS)", round(indicadores['kg_producidos'], 2), "kg", "")
        row += 2

        # Sección: Financiero
        row = escribir_seccion(ws, row, "💰 Financiero")
        escribir_indicador(ws, row, "Ingresos Totales", formatear_moneda_indicador(indicadores['ingresos_totales']), "$", "")
        row += 1
        escribir_indicador(ws, row, "Costos MO", formatear_moneda_indicador(indicadores['costos_mo']), "$", "")
        row += 1
        escribir_indicador(ws, row, "Costos Insumos", formatear_moneda_indicador(indicadores['costos_insumos']), "$", "")
        row += 1
        escribir_indicador(ws, row, "Costos Totales", formatear_moneda_indicador(indicadores['costos_total']), "$", "")
        row += 1
        escribir_indicador(ws, row, "Costo Total por Hectárea", formatear_moneda_indicador(indicadores['costo_total_por_ha']), "$/ha", "")
        row += 1
        escribir_indicador(ws, row, "Costo por Kg CPS", formatear_moneda_indicador(indicadores['costo_por_kilo']), "$/kg", "")
        row += 1
        escribir_indicador(ws, row, "Precio Venta Promedio", formatear_moneda_indicador(indicadores['precio_venta_promedio']), "$/kg", "")
        row += 1

        margen = indicadores['margen_por_ha']
        escribir_indicador(ws, row, "Margen por Hectárea", formatear_moneda_indicador(margen), "$/ha", "")
        row += 2

        # Sección: Productividad
        row = escribir_seccion(ws, row, "📈 Productividad")
        escribir_indicador(ws, row, "Productividad (kg/ha total)", round(indicadores['productividad'], 2), "kg/ha", "60-80")
        row += 1
        escribir_indicador(ws, row, "Rendimiento (kg/ha productivo)", round(indicadores['rendimiento'], 2), "kg/ha", "70-100")
        row += 2

        # Sección: Nuevos indicadores con unidades
        row = escribir_seccion(ws, row, "📐 Indicadores por Unidad")
        escribir_indicador(ws, row, "Costo Insumos por kg CPS", formatear_moneda_indicador(indicadores['costo_insumos_por_kg_cps']), "$/kg", "")
        row += 1
        escribir_indicador(ws, row, "Costo MO por kg CPS", formatear_moneda_indicador(indicadores['costo_mo_por_kg_cps']), "$/kg", "")
        row += 1
        escribir_indicador(ws, row, "Insumos por Hectárea", round(indicadores['insumos_por_ha'], 2), "kg eq./ha", "")
        row += 1
        escribir_indicador(ws, row, "Eficiencia de Insumos", round(indicadores['eficiencia_insumos'], 2), "kg CPS/kg ins.", "")
        row += 2

        # ─── Tabla de conversión de unidades ───
        row = escribir_seccion(ws, row, "📏 Tabla de Conversión de Unidades")
        # Headers
        conv_headers = ["Cantidad", "Unidad", "Convierte a", "Factor"]
        for col_idx, h in enumerate(conv_headers, 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = Font(bold=True, size=10, name="Calibri", color="FFFFFF")
            cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        row += 1

        conversiones = [
            (500, "g", "0.5 kg", "×0.001"),
            (2, "kg", "2 kg", "×1"),
            (1000, "mL", "1 L", "×0.001"),
            (4, "L", "4 L", "×1"),
            (3, "bulto", "150 kg", "×50"),
        ]
        for cant, uni, conv, factor in conversiones:
            ws.cell(row=row, column=1, value=cant).font = FONT_NORMAL
            ws.cell(row=row, column=1).alignment = RIGHT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=2, value=uni).font = FONT_NORMAL
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=3, value=conv).font = FONT_NORMAL
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=3).border = THIN_BORDER
            ws.cell(row=row, column=4, value=factor).font = FONT_NORMAL
            ws.cell(row=row, column=4).alignment = CENTER
            ws.cell(row=row, column=4).border = THIN_BORDER
            row += 1
        row += 2

        # ─── Gráfico 1: MO vs Insumos por Hectárea ───
        chart_start = row + 1
        try:
            ws.cell(row=chart_start, column=1, value="Tipo")
            ws.cell(row=chart_start, column=2, value="Costo por ha ($)")

            ws.cell(row=chart_start + 1, column=1, value="Mano de Obra")
            ws.cell(row=chart_start + 1, column=2, value=round(indicadores['costo_mo_por_ha'], 2))

            ws.cell(row=chart_start + 2, column=1, value="Insumos")
            ws.cell(row=chart_start + 2, column=2, value=round(indicadores['costo_insumos_por_ha'], 2))

            chart1 = BarChart()
            chart1.type = "col"
            chart1.title = "Costo MO vs Insumos por Hectárea"
            chart1.y_axis.title = "Costo ($/ha)"
            chart1.style = 10
            chart1.width = 16
            chart1.height = 10

            data1 = Reference(ws, min_col=2, min_row=chart_start, max_row=chart_start + 2)
            cats1 = Reference(ws, min_col=1, min_row=chart_start + 1, max_row=chart_start + 2)
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(cats1)

            chart1.series[0].graphicalProperties.solidFill = "1565C0"

            ws.add_chart(chart1, f"A{chart_start + 4}")
        except Exception as e:
            logger.warning(f"No se pudo generar gráfico MO vs Insumos: {e}")

        # ─── Gráfico 2: Productividad vs Rendimiento ───
        try:
            chart2_start = chart_start
            ws.cell(row=chart2_start, column=4, value="Métrica")
            ws.cell(row=chart2_start, column=5, value="kg/ha")

            ws.cell(row=chart2_start + 1, column=4, value="Productividad")
            ws.cell(row=chart2_start + 1, column=5, value=round(indicadores['productividad'], 2))

            ws.cell(row=chart2_start + 2, column=4, value="Rendimiento")
            ws.cell(row=chart2_start + 2, column=5, value=round(indicadores['rendimiento'], 2))

            chart2 = BarChart()
            chart2.type = "col"
            chart2.title = "Productividad vs Rendimiento"
            chart2.y_axis.title = "kg/ha"
            chart2.style = 10
            chart2.width = 16
            chart2.height = 10

            data2 = Reference(ws, min_col=5, min_row=chart2_start, max_row=chart2_start + 2)
            cats2 = Reference(ws, min_col=4, min_row=chart2_start + 1, max_row=chart2_start + 2)
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)

            chart2.series[0].graphicalProperties.solidFill = "2E7D32"

            ws.add_chart(chart2, f"D{chart2_start + 4}")
        except Exception as e:
            logger.warning(f"No se pudo generar gráfico de productividad: {e}")

        logger.info(f"Hoja 'Indicadores' generada para finca {finca_id}")

    # ------------------------------------------------------------------
    # Gráficos de tendencia (MEJORA 4)
    # ------------------------------------------------------------------

    def _generar_hoja_graficos(self, db, finca_id: int, ws):
        """
        Genera 3 gráficos de tendencia en la hoja de gráficos:
        1. BarChart: Ingresos vs Egresos por año (2023-2025)
        2. PieChart: Distribución de costos por categoría
        3. LineChart: Evolución del costo total por mes
        """
        conn = db.get_conn()
        try:
            # =============================================================
            # 1. BARCHART: Ingresos vs Egresos por año
            # =============================================================
            ws['A1'] = 'Año'
            ws['B1'] = 'Ingresos'
            ws['C1'] = 'Egresos'

            for i, year in enumerate([2023, 2024, 2025]):
                row = 2 + i
                ws.cell(row=row, column=1, value=year)

                ing = conn.execute(
                    "SELECT COALESCE(SUM(valor_total), 0) FROM transacciones "
                    "WHERE finca_id = ? AND categoria LIKE 'ingreso_%' AND fecha LIKE ?",
                    (finca_id, f"{year}%")
                ).fetchone()[0]
                ws.cell(row=row, column=2, value=ing or 0)

                egr = conn.execute(
                    "SELECT COALESCE(SUM(valor_total), 0) FROM transacciones "
                    "WHERE finca_id = ? AND categoria NOT LIKE 'ingreso_%' AND fecha LIKE ?",
                    (finca_id, f"{year}%")
                ).fetchone()[0]
                ws.cell(row=row, column=3, value=egr or 0)

            chart1 = BarChart()
            chart1.type = "col"
            chart1.title = "Ingresos vs Egresos por Año"
            chart1.y_axis.title = "Valor ($)"
            chart1.style = 10

            data1 = Reference(ws, min_col=1, min_row=1, max_col=3, max_row=4)
            cats1 = Reference(ws, min_col=1, min_row=2, max_row=4)
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(cats1)

            # Colores: Ingresos verde (#2E7D32), Egresos rojo (#C62828)
            chart1.series[0].graphicalProperties.solidFill = "2E7D32"
            chart1.series[1].graphicalProperties.solidFill = "C62828"

            ws.add_chart(chart1, "E1")

            # =============================================================
            # 2. PIECHART: Distribución de costos por categoría
            # =============================================================
            ws['A10'] = 'Categoría'
            ws['B10'] = 'Total'

            categorias_costo = [
                ("Instalación", ["instalacion_mo", "instalacion_insumos"]),
                ("Arvenses", ["arvenses_mo", "arvenses_insumos"]),
                ("Fertilización", ["fertilizacion_mo", "fertilizacion_insumos"]),
                ("Fitosanitario", ["fitosanitario_mo", "fitosanitario_insumos"]),
                ("Sombrio", ["sombrio_mo", "sombrio_insumos"]),
                ("Otras Labores", ["otras_labores_mo", "otras_labores_insumos"]),
                ("Recolección", ["recoleccion"]),
                ("Beneficio", ["beneficio"]),
                ("Administrativo", ["administrativo"]),
            ]

            pie_palette = [
                "2E7D32", "1565C0", "F9A825", "E65100",
                "6A1B9A", "00838F", "C62828", "4E342E", "558B2F",
            ]

            for i, (cat_name, cat_db_names) in enumerate(categorias_costo):
                row = 11 + i
                ws.cell(row=row, column=1, value=cat_name)

                placeholders = ",".join("?" for _ in cat_db_names)
                total = conn.execute(
                    f"SELECT COALESCE(SUM(valor_total), 0) FROM transacciones "
                    f"WHERE finca_id = ? AND categoria IN ({placeholders})",
                    (finca_id, *cat_db_names)
                ).fetchone()[0]
                ws.cell(row=row, column=2, value=total or 0)

            chart2 = PieChart()
            chart2.title = "Distribución de Costos por Categoría"
            chart2.style = 10

            data2 = Reference(ws, min_col=2, min_row=10, max_row=10 + len(categorias_costo))
            cats2 = Reference(ws, min_col=1, min_row=11, max_row=10 + len(categorias_costo))
            chart2.add_data(data2, titles_from_data=True)
            chart2.set_categories(cats2)

            # Data labels con porcentaje y nombre de categoría
            chart2.dataLabels = DataLabelList()
            chart2.dataLabels.showPercent = True
            chart2.dataLabels.showCatName = True

            # Colores para cada sector del pie
            for i in range(len(categorias_costo)):
                pt = DataPoint(idx=i)
                pt.graphicalProperties.solidFill = pie_palette[i]
                chart2.series[0].data_points.append(pt)

            ws.add_chart(chart2, "E17")

            # =============================================================
            # 3. LINECHART: Costo total por mes (todas las categorías de egreso)
            # =============================================================
            ws['A20'] = 'Mes'
            ws['B20'] = 'Costo Total'

            meses = [
                "Ene", "Feb", "Mar", "Abr", "May", "Jun",
                "Jul", "Ago", "Sep", "Oct", "Nov", "Dic",
            ]

            for i, mes_nombre in enumerate(meses):
                row = 21 + i
                ws.cell(row=row, column=1, value=mes_nombre)
                mes_num = f"{i + 1:02d}"

                total = conn.execute(
                    "SELECT COALESCE(SUM(valor_total), 0) FROM transacciones "
                    "WHERE finca_id = ? AND categoria NOT LIKE 'ingreso_%' "
                    "AND substr(fecha, 6, 2) = ?",
                    (finca_id, mes_num)
                ).fetchone()[0]
                ws.cell(row=row, column=2, value=total or 0)

            chart3 = LineChart()
            chart3.title = "Costo Total por Mes"
            chart3.y_axis.title = "Costo ($)"
            chart3.x_axis.title = "Mes"
            chart3.style = 10

            data3 = Reference(ws, min_col=2, min_row=20, max_row=32)
            cats3 = Reference(ws, min_col=1, min_row=21, max_row=32)
            chart3.add_data(data3, titles_from_data=True)
            chart3.set_categories(cats3)

            # Color de la línea: azul
            chart3.series[0].graphicalProperties.line.solidFill = "1565C0"
            chart3.series[0].graphicalProperties.line.width = 25000  # 2.5pt en EMUs

            ws.add_chart(chart3, "E33")

            logger.info("Hoja 'Gráficos' generada con 3 gráficos (Bar, Pie, Line)")

        finally:
            conn.close()

    # ------------------------------------------------------------------
    # Utilidades
    # ------------------------------------------------------------------

    def _poner_fecha(self, ws, fila: int, col: int, fecha_str: str):
        """Pone una fecha en una celda, intentando convertir a datetime."""
        if not fecha_str:
            ws.cell(row=fila, column=col, value="")
            return

        try:
            from datetime import datetime as dt
            # Intentar varios formatos
            for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
                try:
                    fecha_dt = dt.strptime(fecha_str, fmt)
                    cell = ws.cell(row=fila, column=col, value=fecha_dt)
                    cell.number_format = "DD/MM/YYYY"
                    return
                except ValueError:
                    continue
            # Si no funciona ningún formato, poner como string
            ws.cell(row=fila, column=col, value=fecha_str)
        except Exception as e:
            logger.warning(f"Error al parsear fecha '{fecha_str}': {e}")
            ws.cell(row=fila, column=col, value=fecha_str)
