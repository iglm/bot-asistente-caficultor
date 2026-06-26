"""
API Backend completo para Mini App — 100% funcionalidad del bot.
Asistente de Costos para Caficultores ☕
"""
import os, sys, json, io, tempfile, traceback
from datetime import datetime, timedelta
from typing import Optional
from fastapi import FastAPI, HTTPException, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', '..'))
from database import Database
from config import (
    CATEGORIAS_PADRE, CATEGORIAS_SIMPLE, TIPOS_CAFE_LIST,
    UNIDADES_INSUMO, UNIDADES_INSUMO_LABELS,
    FNC_INDICADORES, PRESUPUESTO_PORCENTAJES,
    ASESORIA_EMAIL, ASESOR_NOMBRE, ASESOR_ASESOR,
    AVISO_LEGAL, ADMIN_IDS,
)

app = FastAPI(title="Asistente de Costos API", version="2.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True,
                   allow_methods=["*"], allow_headers=["*"])
db = Database()

# ─── Pydantic Models ───

class TransaccionCreate(BaseModel):
    finca_id: int
    lote_id: int = 0
    categoria: str
    fecha: str
    labor: str = ""
    producto: str = ""
    cantidad: float = 0
    unidad: str = ""
    valor_unitario: float = 0
    valor_total: float = 0

class FincaCreate(BaseModel):
    user_id: int
    nombre: str
    region: str = ""
    departamento: str = ""

class LoteCreate(BaseModel):
    finca_id: int
    nombre: str
    area_hectareas: float = 0
    num_arboles: int = 0
    variedad: str = ""
    fecha_siembra: str = ""

class PresupuestoCreate(BaseModel):
    finca_id: int
    anio: int
    datos: dict  # {categoria: monto}

class CostoMOCreate(BaseModel):
    finca_id: int
    lote_id: int = 0
    categoria: str
    fecha: str
    labor: str = ""
    cantidad: float = 0
    valor_unitario: float = 0
    valor_total: float = 0

class CostoInsumoCreate(BaseModel):
    finca_id: int
    lote_id: int = 0
    categoria: str
    fecha: str
    producto: str = ""
    cantidad: float = 0
    unidad: str = ""
    valor_unitario: float = 0
    valor_total: float = 0

# ═══════════════════════════════════════════
# HEALTH
# ═══════════════════════════════════════════

@app.get("/api/health")
def health_check():
    return {"status": "ok", "timestamp": datetime.now().isoformat(),
            "version": "2.0", "app": "Asistente de Costos para Caficultores"}

# ═══════════════════════════════════════════
# CONFIG / CONSTANTES
# ═══════════════════════════════════════════

@app.get("/api/config")
def get_config():
    return {
        "app_name": "Asistente de Costos para Caficultores ☕",
        "version": "2.0.0",
        "developer": ASESOR_NOMBRE,
        "advisor": ASESOR_ASESOR,
        "contact": ASESORIA_EMAIL,
        "aviso_legal": AVISO_LEGAL,
        "categorias_padre": {k: v["nombre"] for k, v in CATEGORIAS_PADRE.items()},
        "categorias_simple": {k: v["nombre"] for k, v in CATEGORIAS_SIMPLE.items()},
        "tipos_cafe": TIPOS_CAFE_LIST,
        "unidades_insumo": {k: UNIDADES_INSUMO_LABELS.get(k, k) for k in UNIDADES_INSUMO},
        "fnc_indicadores": FNC_INDICADORES,
        "presupuesto_porcentajes": PRESUPUESTO_PORCENTAJES,
    }

# ═══════════════════════════════════════════
# USUARIOS / AUTENTICACIÓN
# ═══════════════════════════════════════════

@app.get("/api/usuarios/{user_id}")
def get_usuario(user_id: int):
    try:
        user = db.get_user(user_id)
        if not user:
            db.upsert_user(user_id, "")
            user = db.get_user(user_id)
        if not user:
            raise HTTPException(404, "Usuario no encontrado")
        result = dict(user)
        # Asegurar campo acepto_terminos
        if "acepto_terminos" not in result:
            result["acepto_terminos"] = 0
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error interno: {str(e)}")

@app.post("/api/usuarios/{user_id}/aceptar-terminos")
def aceptar_terminos(user_id: int):
    ok = db.aceptar_terminos(user_id)
    if not ok:
        raise HTTPException(404, "Usuario no encontrado")
    return {"status": "ok", "message": "Términos aceptados"}

@app.post("/api/auth/register")
def register_user(user_id: int = Form(...), username: str = Form("")):
    es_nuevo = db.register_user(user_id, username)
    return {"status": "ok", "es_nuevo": es_nuevo, "message": "Usuario registrado"}

# ═══════════════════════════════════════════
# FINCAS — CRUD COMPLETO
# ═══════════════════════════════════════════

@app.get("/api/fincas/user/{user_id}")
def listar_fincas(user_id: int):
    fincas = db.get_fincas(user_id)
    return {"fincas": [dict(f) for f in fincas]}

@app.get("/api/fincas/{finca_id}")
def get_finca(finca_id: int):
    f = db.get_finca(finca_id)
    if not f:
        raise HTTPException(404, "Finca no encontrada")
    return dict(f)

@app.post("/api/fincas")
def crear_finca(data: FincaCreate):
    fid = db.create_finca(data.user_id, data.nombre, data.region, data.departamento)
    return {"status": "created", "id": fid, "message": f"Finca '{data.nombre}' creada"}

@app.put("/api/fincas/{finca_id}")
def editar_finca(finca_id: int, nombre: str = Form(""), region: str = Form(""), departamento: str = Form("")):
    import sqlite3
    conn = db.get_conn()
    try:
        updates = []
        params = []
        if nombre: updates.append("nombre=?"); params.append(nombre)
        if region: updates.append("region=?"); params.append(region)
        if departamento: updates.append("departamento=?"); params.append(departamento)
        if not updates:
            raise HTTPException(400, "Sin datos para actualizar")
        params.append(finca_id)
        conn.execute(f"UPDATE fincas SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
        return {"status": "updated", "message": "Finca actualizada"}
    finally:
        conn.close()

@app.delete("/api/fincas/{finca_id}")
def eliminar_finca(finca_id: int):
    f = db.get_finca(finca_id)
    if not f:
        raise HTTPException(404, "Finca no encontrada")
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM transacciones WHERE finca_id=?", (finca_id,))
        conn.execute("DELETE FROM lotes WHERE finca_id=?", (finca_id,))
        conn.execute("DELETE FROM presupuestos WHERE finca_id=?", (finca_id,))
        conn.execute("DELETE FROM fincas WHERE id=?", (finca_id,))
        conn.commit()
        return {"status": "deleted", "message": f"Finca '{f['nombre']}' eliminada"}
    finally:
        conn.close()

# ═══════════════════════════════════════════
# LOTES — CRUD COMPLETO
# ═══════════════════════════════════════════

@app.get("/api/lotes/{finca_id}")
def listar_lotes(finca_id: int):
    lotes = db.get_lotes(finca_id)
    return {"lotes": [dict(l) for l in lotes]}

@app.get("/api/lotes/detalle/{lote_id}")
def get_lote(lote_id: int):
    l = db.get_lote_by_id(lote_id)
    if not l:
        raise HTTPException(404, "Lote no encontrado")
    return dict(l)

@app.post("/api/lotes")
def crear_lote(data: LoteCreate):
    lid = db.create_lote(
        data.finca_id, data.nombre, data.area_hectareas,
        data.num_arboles, data.variedad, data.fecha_siembra
    )
    return {"status": "created", "id": lid, "message": f"Lote '{data.nombre}' creado"}

@app.put("/api/lotes/{lote_id}")
def editar_lote(lote_id: int, nombre: str = Form(""), area_hectareas: float = Form(0),
                num_arboles: int = Form(0), variedad: str = Form(""), fecha_siembra: str = Form("")):
    conn = db.get_conn()
    try:
        updates = []; params = []
        if nombre: updates.append("nombre=?"); params.append(nombre)
        if area_hectareas: updates.append("area_hectareas=?"); params.append(area_hectareas)
        if num_arboles: updates.append("num_arboles=?"); params.append(num_arboles)
        if variedad: updates.append("variedad=?"); params.append(variedad)
        if fecha_siembra: updates.append("fecha_siembra=?"); params.append(fecha_siembra)
        if not updates: raise HTTPException(400, "Sin datos")
        params.append(lote_id)
        conn.execute(f"UPDATE lotes SET {', '.join(updates)} WHERE id=?", params)
        conn.commit()
        return {"status": "updated"}
    finally:
        conn.close()

@app.delete("/api/lotes/{lote_id}")
def eliminar_lote(lote_id: int):
    l = db.get_lote_by_id(lote_id)
    if not l:
        raise HTTPException(404, "Lote no encontrado")
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM transacciones WHERE lote_id=?", (lote_id,))
        conn.execute("DELETE FROM lotes WHERE id=?", (lote_id,))
        conn.commit()
        return {"status": "deleted", "message": f"Lote '{l['nombre']}' eliminado"}
    finally:
        conn.close()

# ═══════════════════════════════════════════
# TRANSACCIONES / INGRESOS
# ═══════════════════════════════════════════

@app.post("/api/transacciones")
def crear_transaccion(t: TransaccionCreate):
    tid = db.insert_transaccion(
        finca_id=t.finca_id, categoria=t.categoria, fecha=t.fecha,
        labor=t.labor, producto=t.producto, cantidad=t.cantidad,
        unidad=t.unidad, valor_unitario=t.valor_unitario,
        valor_total=t.valor_total, lote_id=t.lote_id,
    )
    return {"status": "created", "id": tid}

@app.get("/api/transacciones/{finca_id}")
def listar_transacciones(
    finca_id: int,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    categoria: Optional[str] = None,
    limit: int = 200,
):
    if fecha_inicio and fecha_fin:
        tx_list = db.get_transacciones_por_periodo(finca_id, fecha_inicio, fecha_fin)
    else:
        tx_list = db.get_all_transacciones(finca_id)
    if categoria:
        tx_list = [t for t in tx_list if t["categoria"] == categoria]
    return {"transacciones": [dict(t) for t in tx_list][-limit:], "total": len(tx_list)}

@app.get("/api/transacciones/detalle/{tx_id}")
def detalle_transaccion(tx_id: int):
    conn = db.get_conn()
    try:
        row = conn.execute("SELECT * FROM transacciones WHERE id=?", (tx_id,)).fetchone()
        return dict(row) if row else raise_error(404, "No encontrada")
    finally:
        conn.close()

@app.delete("/api/transacciones/{tx_id}")
def eliminar_transaccion(tx_id: int):
    conn = db.get_conn()
    try:
        conn.execute("DELETE FROM transacciones WHERE id=?", (tx_id,))
        conn.commit()
        return {"status": "deleted"}
    finally:
        conn.close()

@app.get("/api/ingresos/tipos")
def tipos_cafe():
    return {"tipos": TIPOS_CAFE_LIST}

# ═══════════════════════════════════════════
# COSTOS — CATEGORÍAS, UNIDADES, REGISTRO
# ═══════════════════════════════════════════

@app.get("/api/costos/categorias")
def costos_categorias():
    return {
        "categorias_padre": {k: v["nombre"] for k, v in CATEGORIAS_PADRE.items()},
        "categorias_simple": {k: v["nombre"] for k, v in CATEGORIAS_SIMPLE.items()},
    }

@app.get("/api/costos/unidades")
def costos_unidades():
    return {"unidades": {k: UNIDADES_INSUMO_LABELS.get(k, k) for k in UNIDADES_INSUMO}}

@app.post("/api/costos/mo")
def registrar_costo_mo(data: CostoMOCreate):
    # La categoría MO se registra directamente
    tid = db.insert_transaccion(
        finca_id=data.finca_id, lote_id=data.lote_id,
        categoria=data.categoria, fecha=data.fecha,
        labor=data.labor, producto=data.labor,
        cantidad=data.cantidad, unidad="jornal",
        valor_unitario=data.valor_unitario,
        valor_total=data.valor_total,
    )
    return {"status": "created", "id": tid}

@app.post("/api/costos/insumos")
def registrar_costo_insumo(data: CostoInsumoCreate):
    tid = db.insert_transaccion(
        finca_id=data.finca_id, lote_id=data.lote_id,
        categoria=data.categoria, fecha=data.fecha,
        labor=data.producto, producto=data.producto,
        cantidad=data.cantidad, unidad=data.unidad,
        valor_unitario=data.valor_unitario,
        valor_total=data.valor_total,
    )
    return {"status": "created", "id": tid}

# ═══════════════════════════════════════════
# RESUMEN FINANCIERO
# ═══════════════════════════════════════════

@app.get("/api/resumen/{finca_id}")
def get_resumen(finca_id: int):
    resumen = db.get_resumen_finca(finca_id)
    return resumen

@app.get("/api/resumen/{finca_id}/periodo")
def get_resumen_periodo(
    finca_id: int,
    fecha_inicio: str = Query(...),
    fecha_fin: str = Query(...),
):
    resumen = db.get_resumen_por_periodo(finca_id, fecha_inicio, fecha_fin)
    return resumen

@app.get("/api/resumen/{finca_id}/semanal")
def get_resumen_semanal(finca_id: int, anio: int, semana: int):
    return db.get_resumen_semanal(finca_id, anio, semana)

@app.get("/api/resumen/{finca_id}/mensual")
def get_resumen_mensual(finca_id: int, anio: int, mes: int):
    return db.get_resumen_mensual(finca_id, anio, mes)

@app.get("/api/resumen/{finca_id}/anual")
def get_resumen_anual(finca_id: int, anio: int):
    return db.get_resumen_anual(finca_id, anio)

# ═══════════════════════════════════════════
# INDICADORES TÉCNICOS
# ═══════════════════════════════════════════

@app.get("/api/indicadores/{finca_id}")
def get_indicadores(finca_id: int):
    ind = db.get_indicadores_tecnicos(finca_id)
    return ind if ind else raise_error(404, "Sin datos")

@app.get("/api/indicadores/{finca_id}/mo")
def get_indicadores_mo(finca_id: int):
    ind = db.get_indicadores_tecnicos(finca_id)
    if not ind:
        raise HTTPException(404, "Sin datos")
    return {
        "total_jornales": ind["total_jornales"],
        "costos_mo": ind["costos_mo"],
        "jornales_por_ha": ind["jornales_por_ha"],
        "costo_mo_por_ha": ind["costo_mo_por_ha"],
        "costo_mo_por_kg_cps": ind["costo_mo_por_kg_cps"],
        "eficiencia_mo": ind["eficiencia_mo"],
    }

@app.get("/api/indicadores/{finca_id}/insumos")
def get_indicadores_insumos(finca_id: int):
    ind = db.get_indicadores_tecnicos(finca_id)
    if not ind:
        raise HTTPException(404, "Sin datos")
    return {
        "costos_insumos": ind["costos_insumos"],
        "costo_insumos_por_ha": ind["costo_insumos_por_ha"],
        "costo_insumos_por_kg_cps": ind["costo_insumos_por_kg_cps"],
        "insumos_por_ha": ind["insumos_por_ha"],
        "insumos_total_kg": ind["insumos_total_kg"],
        "insumos_total_litros": ind["insumos_total_litros"],
        "eficiencia_insumos": ind["eficiencia_insumos"],
    }

@app.get("/api/indicadores/{finca_id}/financiero")
def get_indicadores_financiero(finca_id: int):
    ind = db.get_indicadores_tecnicos(finca_id)
    if not ind:
        raise HTTPException(404, "Sin datos")
    return {
        "ingresos_totales": ind["ingresos_totales"],
        "costos_total": ind["costos_total"],
        "margen_por_ha": ind["margen_por_ha"],
        "precio_venta_promedio": ind["precio_venta_promedio"],
        "costo_por_kilo": ind["costo_por_kilo"],
        "productividad": ind["productividad"],
    }

@app.get("/api/indicadores/referencia-fnc")
def referencia_fnc():
    return FNC_INDICADORES

# ═══════════════════════════════════════════
# PRESUPUESTOS
# ═══════════════════════════════════════════

@app.post("/api/presupuesto/{finca_id}")
def guardar_presupuesto(finca_id: int, data: PresupuestoCreate):
    db.guardar_presupuesto(finca_id, data.anio, data.datos)
    return {"status": "saved", "message": f"Presupuesto {data.anio} guardado"}

@app.get("/api/presupuesto/{finca_id}")
def get_presupuesto(finca_id: int, anio: Optional[int] = None):
    if anio:
        presupuesto = db.get_presupuesto(finca_id, anio)
        return {"presupuesto": [dict(p) for p in presupuesto], "anio": anio}
    anios = db.get_presupuesto_anios(finca_id)
    return {"anios": anios}

@app.get("/api/presupuesto/{finca_id}/anios")
def get_presupuesto_anios(finca_id: int):
    anios = db.get_presupuesto_anios(finca_id)
    return {"anios": anios}

@app.get("/api/presupuesto/{finca_id}/ejecucion/{anio}")
def get_ejecucion_presupuesto(finca_id: int, anio: int):
    ejecucion = db.get_ejecucion_presupuesto(finca_id, anio)
    return ejecucion

@app.get("/api/presupuesto/{finca_id}/sugerido")
def get_presupuesto_sugerido(finca_id: int):
    """Calcula montos sugeridos basados en área total y porcentajes FNC."""
    conn = db.get_conn()
    try:
        area = conn.execute(
            "SELECT SUM(area_hectareas) FROM lotes WHERE finca_id=?", (finca_id,)
        ).fetchone()[0] or 1.0
    finally:
        conn.close()

    costo_base_ha = FNC_INDICADORES["costo_ha"]  # ~16.3M/ha
    costo_total = costo_base_ha * area

    sugerido = {}
    for rubro, pct in PRESUPUESTO_PORCENTAJES.items():
        sugerido[rubro] = round(costo_total * pct, -3)  # redondear a miles
    return {"sugerido": sugerido, "area": area, "costo_total_referencia": costo_total}

@app.delete("/api/presupuesto/{finca_id}")
def eliminar_presupuesto(finca_id: int, anio: int = Query(...)):
    db.delete_presupuesto(finca_id, anio)
    return {"status": "deleted"}

# ═══════════════════════════════════════════
# ASESORÍA
# ═══════════════════════════════════════════

@app.get("/api/asesoria/interpretar/{finca_id}")
def asesoria_interpretar(finca_id: int):
    ind = db.get_indicadores_tecnicos(finca_id)
    if not ind:
        raise HTTPException(404, "Sin datos para interpretar")

    analisis = []
    # Productividad
    if ind["productividad"] < FNC_INDICADORES["productividad_ha"]:
        analisis.append({
            "tipo": "warning",
            "indicador": "Productividad",
            "valor": f"{ind['productividad']:.1f} kg/ha",
            "referencia": f"{FNC_INDICADORES['productividad_ha']:,} kg/ha",
            "mensaje": "Tu productividad está por debajo del promedio nacional. "
                       "Considerá renovar lotes viejos y mejorar la fertilización."
        })
    else:
        analisis.append({
            "tipo": "success",
            "indicador": "Productividad",
            "valor": f"{ind['productividad']:.1f} kg/ha",
            "referencia": f"{FNC_INDICADORES['productividad_ha']:,} kg/ha",
            "mensaje": "¡Buena productividad! Estás por encima del promedio nacional."
        })

    # Costo por kilo
    if ind["costo_por_kilo"] > FNC_INDICADORES["costo_produccion_kilo"]:
        analisis.append({
            "tipo": "danger",
            "indicador": "Costo de producción",
            "valor": f"${ind['costo_por_kilo']:,.0f}/kg",
            "referencia": f"${FNC_INDICADORES['costo_produccion_kilo']:,}/kg",
            "mensaje": "Tus costos por kilo superan el promedio nacional. "
                       "Revisá la eficiencia de tus labores y el uso de insumos."
        })
    else:
        analisis.append({
            "tipo": "success",
            "indicador": "Costo de producción",
            "valor": f"${ind['costo_por_kilo']:,.0f}/kg",
            "referencia": f"${FNC_INDICADORES['costo_produccion_kilo']:,}/kg",
            "mensaje": "Tus costos están por debajo del promedio. Bien manejado."
        })

    # Margen
    margen = (ind["ingresos_totales"] - ind["costos_total"])
    if margen < 0:
        analisis.append({
            "tipo": "danger",
            "indicador": "Margen neto",
            "valor": f"${margen:,.0f}",
            "referencia": "Positivo",
            "mensaje": "⚠️ Estás operando con pérdidas. Es urgente reducir costos "
                       "o mejorar el precio de venta."
        })
    else:
        analisis.append({
            "tipo": "success",
            "indicador": "Margen neto",
            "valor": f"${margen:,.0f}",
            "referencia": "Positivo",
            "mensaje": "Tu finca genera ganancias. Seguí optimizando."
        })

    # Eficiencia MO
    if ind["eficiencia_mo"] < 20 and ind["total_jornales"] > 0:
        analisis.append({
            "tipo": "warning",
            "indicador": "Eficiencia MO",
            "valor": f"{ind['eficiencia_mo']:.1f} kg/jornal",
            "referencia": "> 20 kg/jornal",
            "mensaje": "La eficiencia de mano de obra es baja. Considerá "
                       "capacitación y herramientas que mejoren la productividad."
        })

    # Comparativa de precio
    if ind["precio_venta_promedio"] < FNC_INDICADORES["precio_venta_promedio"] and ind["kg_producidos"] > 0:
        analisis.append({
            "tipo": "warning",
            "indicador": "Precio de venta",
            "valor": f"${ind['precio_venta_promedio']:,.0f}/kg",
            "referencia": f"${FNC_INDICADORES['precio_venta_promedio']:,}/kg",
            "mensaje": "Tu precio promedio está por debajo del nacional. "
                       "Explorá canales de comercialización con mejor precio."
        })

    return {"analisis": analisis, "finca_id": finca_id}

@app.get("/api/asesoria/sugerencias")
def asesoria_sugerencias():
    return {
        "sugerencias": [
            {
                "icono": "📊",
                "titulo": "Llevá registros detallados",
                "descripcion": "Registrá cada ingreso y egreso para tener visibilidad completa de tus costos."
            },
            {
                "icono": "🧪",
                "titulo": "Optimizá la fertilización",
                "descripcion": "Basá tus planes de fertilización en análisis de suelo. Aplicá la dosis justa."
            },
            {
                "icono": "🌱",
                "titulo": "Renová lotes viejos",
                "descripcion": "Los cafetales mayores a 8 años pierden productividad. Planificá la renovación."
            },
            {
                "icono": "💰",
                "titulo": "Buscá mejores precios",
                "descripcion": "Cotizá en diferentes compradores. El café de especialidad paga mejores precios."
            },
            {
                "icono": "📋",
                "titulo": "Usá presupuestos",
                "descripcion": "Planificá tus gastos anuales. El presupuesto te ayuda a controlar costos."
            },
        ]
    }

@app.get("/api/asesoria/plan")
def asesoria_plan():
    return {
        "plan": {
            "corto_plazo": [
                "📝 Registrar todas las transacciones diariamente",
                "📊 Revisar el resumen financiero cada semana",
                "💰 Comparar precios de venta disponibles",
            ],
            "mediano_plazo": [
                "🌱 Implementar plan de fertilización basado en suelos",
                "📋 Crear presupuesto anual y hacer seguimiento",
                "👨‍🌾 Capacitar al personal en buenas prácticas",
            ],
            "largo_plazo": [
                "🔄 Plan de renovación de cafetales por etapas",
                "🏆 Certificación de café de especialidad",
                "📈 Diversificar canales de comercialización",
            ]
        }
    }

@app.get("/api/asesoria/contacto")
def asesoria_contacto():
    return {
        "nombre": ASESOR_NOMBRE,
        "asesor": ASESOR_ASESOR,
        "email": ASESORIA_EMAIL,
        "mensaje": "Contactanos para recibir asesoría personalizada gratuita."
    }

# ═══════════════════════════════════════════
# EXPORT / IMPORT EXCEL
# ═══════════════════════════════════════════

@app.get("/api/excel/{finca_id}")
def exportar_excel(finca_id: int):
    """Genera y descarga Excel completo con datos de la finca."""
    try:
        from excel_manager import ExcelManager
        output_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "exports", f"finca_{finca_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        ExcelManager.generar_excel(finca_id, db, output_path)
        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=f"costos_finca_{finca_id}.xlsx"
        )
    except Exception as e:
        raise HTTPException(500, f"Error generando Excel: {str(e)}")

@app.get("/api/excel/plantilla")
def descargar_plantilla():
    """Descarga plantilla vacía para importar datos."""
    try:
        from excel_manager import ExcelManager
        output_path = os.path.join(
            os.path.dirname(os.path.dirname(os.path.dirname(__file__))),
            "exports", f"plantilla_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        ExcelManager.generar_plantilla_vacia(output_path)
        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="plantilla_costos.xlsx"
        )
    except Exception as e:
        raise HTTPException(500, f"Error generando plantilla: {str(e)}")

@app.post("/api/excel/importar")
async def importar_excel(file: UploadFile = File(...), user_id: int = Form(...)):
    """Importa datos desde Excel. Lee las hojas y crea registros."""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Formato no soportado. Usá .xlsx")

    try:
        import openpyxl
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))

        resultado = {"importados": [], "errores": [], "fincas_creadas": 0}

        # Hoja: Fincas
        if "Fincas" in wb.sheetnames:
            ws = wb["Fincas"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    nombre, region, depto = str(row[0]), str(row[1] or ""), str(row[2] or "")
                    try:
                        fid = db.create_finca(user_id, nombre, region, depto)
                        resultado["fincas_creadas"] += 1
                        resultado["importados"].append({"tipo": "finca", "nombre": nombre})
                    except Exception as e:
                        resultado["errores"].append({"finca": nombre, "error": str(e)})

        # Hoja: Lotes
        if "Lotes" in wb.sheetnames:
            ws = wb["Lotes"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0] and row[1]:
                    finca_nombre, nombre, area, arboles, variedad, siembra = (
                        str(row[0]), str(row[1]), row[2] or 0, row[3] or 0,
                        str(row[4] or ""), str(row[5] or "")
                    )
                    fincas_u = db.get_fincas(user_id)
                    finca = next((f for f in fincas_u if f["nombre"] == finca_nombre), None)
                    if finca:
                        try:
                            db.create_lote(finca["id"], nombre, float(area), int(arboles), variedad, siembra)
                            resultado["importados"].append({"tipo": "lote", "nombre": nombre})
                        except Exception as e:
                            resultado["errores"].append({"lote": nombre, "error": str(e)})

        wb.close()
        return {"status": "ok", "resultado": resultado}

    except Exception as e:
        raise HTTPException(500, f"Error importando: {str(e)}")

@app.get("/api/excel/preview/{finca_id}")
def preview_export(finca_id: int):
    """Preview de datos a exportar."""
    data = db.get_all_data_for_export(finca_id)
    resumen = db.get_resumen_finca(finca_id)
    return {
        "lotes": len(data.get("lotes", [])),
        "transacciones": sum(len(v) for k, v in data.items() if k != "lotes"),
        "resumen": resumen,
    }

# ═══════════════════════════════════════════
# ADMIN — GESTIÓN DE USUARIOS
# ═══════════════════════════════════════════

@app.get("/api/admin/usuarios")
def admin_listar_usuarios():
    return {
        "todos": [dict(u) for u in db.get_all_users()],
        "pendientes": [dict(u) for u in db.get_pending_users()],
        "aprobados": [dict(u) for u in db.get_approved_users()],
        "rechazados": [dict(u) for u in db.get_rejected_users()],
    }

@app.post("/api/admin/usuarios/aprobar/{user_id}")
def admin_aprobar(user_id: int, admin_id: int = Query(810796748)):
    ok = db.approve_user(user_id, admin_id)
    if not ok:
        raise HTTPException(400, "No se pudo aprobar. Verificá que esté pendiente.")
    return {"status": "approved"}

@app.post("/api/admin/usuarios/rechazar/{user_id}")
def admin_rechazar(user_id: int):
    ok = db.reject_user(user_id)
    if not ok:
        raise HTTPException(400, "No se pudo rechazar")
    return {"status": "rejected"}

@app.post("/api/admin/usuarios/revocar/{user_id}")
def admin_revocar(user_id: int):
    ok = db.revoke_user(user_id)
    if not ok:
        raise HTTPException(400, "No se pudo revocar")
    return {"status": "revoked"}

@app.post("/api/admin/usuarios/reactivar/{user_id}")
def admin_reactivar(user_id: int):
    ok = db.reactivate_user(user_id)
    if not ok:
        raise HTTPException(400, "No se pudo reactivar")
    return {"status": "reactivated"}

@app.delete("/api/admin/borrar-datos/{user_id}")
def admin_borrar_datos(user_id: int):
    resumen = db.delete_all_user_data(user_id)
    return {"status": "deleted", "resumen": resumen}

@app.get("/api/admin/stats")
def admin_stats():
    return {
        "total_usuarios": len(db.get_all_users()),
        "pendientes": len(db.get_pending_users()),
        "aprobados": len(db.get_approved_users()),
        "rechazados": len(db.get_rejected_users()),
    }

# ═══════════════════════════════════════════
# ALERTAS
# ═══════════════════════════════════════════

@app.get("/api/alertas/{finca_id}")
def get_alertas(finca_id: int):
    """Genera alertas automáticas: costo>presupuesto, margen negativo, etc."""
    alertas = []

    # Margen negativo
    resumen = db.get_resumen_finca(finca_id)
    if resumen["margen"] < 0:
        alertas.append({
            "tipo": "danger",
            "titulo": "Margen negativo",
            "mensaje": f"Tu finca tiene un margen negativo de ${abs(resumen['margen']):,.0f}. "
                       "Es urgente revisar costos."
        })

    # Productividad baja vs FNC
    ind = db.get_indicadores_tecnicos(finca_id)
    if ind and ind.get("area_total", 0) > 0:
        if ind["productividad"] < FNC_INDICADORES["productividad_ha"] * 0.7:
            alertas.append({
                "tipo": "warning",
                "titulo": "Productividad baja",
                "mensaje": f"Productividad de {ind['productividad']:.1f} kg/ha "
                           f"vs {FNC_INDICADORES['productividad_ha']:,} kg/ha nacional. "
                           "Considerá renovar o fertilizar."
            })

        if ind["costo_por_kilo"] > FNC_INDICADORES["costo_produccion_kilo"]:
            alertas.append({
                "tipo": "warning",
                "titulo": "Costo de producción alto",
                "mensaje": f"Costo de ${ind['costo_por_kilo']:,.0f}/kg supera "
                           f"el promedio de ${FNC_INDICADORES['costo_produccion_kilo']:,}/kg."
            })

    # Presupuesto vs real
    try:
        anio_actual = datetime.now().year
        ejecucion = db.get_ejecucion_presupuesto(finca_id, anio_actual)
        for cat in ejecucion.get("categorias", []):
            if cat["pct_ejecucion"] > 90 and cat["monto_planificado"] > 0:
                alertas.append({
                    "tipo": "warning",
                    "titulo": f"Categoría {cat['categoria']} al {cat['pct_ejecucion']}% del presupuesto",
                    "mensaje": f"Has usado el {cat['pct_ejecucion']}% del presupuesto en {cat['categoria']}. "
                               f"Plan: ${cat['monto_planificado']:,.0f}, Real: ${cat['monto_ejecutado']:,.0f}"
                })
    except:
        pass

    return {"alertas": alertas, "total": len(alertas)}

# ─── Helper ───

def raise_error(code: int, msg: str):
    raise HTTPException(status_code=code, detail=msg)

# ─── Static Files ───

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
app.mount("/", StaticFiles(directory=BASE_DIR, html=True), name="mini-app")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8080)
