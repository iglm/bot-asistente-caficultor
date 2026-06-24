"""
Handler de Importar Excel — Recibe archivo .xlsx, parsea, muestra preview y confirma.
"""
import os
import logging
import tempfile
from datetime import datetime
from aiogram import Router, types, F
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import BufferedInputFile

from database import Database
from config import CATEGORIAS

from utils import boton_menu, botones_menu_cancelar, agregar_boton_menu

logger = logging.getLogger(__name__)


class ImportExcelState(StatesGroup):
    """Estados para el flujo de importación."""
    esperando_archivo = State()
    preview_mostrado = State()
    confirmado = State()


def get_importar_router(db: Database) -> Router:
    """Router de importación de Excel."""
    router = Router()

    # Mapeo de nombres de columnas esperados por hoja
    HOJAS_ESPERADAS = {
        "Fincas": ["nombre", "region", "departamento"],
        "Lotes": ["finca_nombre", "nombre", "area_hectareas", "num_arboles", "variedad", "fecha_siembra"],
        "Ingresos": ["finca_nombre", "tipo", "fecha", "cantidad", "valor_total"],
        "Costos_MO": ["finca_nombre", "lote_nombre", "categoria", "fecha", "labor", "cantidad", "valor_unitario", "valor_total"],
        "Costos_Insumos": ["finca_nombre", "lote_nombre", "categoria", "fecha", "producto", "cantidad", "unidad", "valor_unitario", "valor_total"],
    }

    CATEGORIAS_MO_VALIDAS = [
        "instalacion_mo", "arvenses_mo", "fertilizacion_mo",
        "fitosanitario_mo", "sombrio_mo", "otras_labores_mo",
        "recoleccion", "beneficio", "administrativo",
    ]
    CATEGORIAS_INSUMOS_VALIDAS = [
        "instalacion_insumos", "arvenses_insumos", "fertilizacion_insumos",
        "fitosanitario_insumos", "sombrio_insumos", "otras_labores_insumos",
    ]
    CATEGORIAS_INGRESO_VALIDAS = ["ingreso_cps", "ingreso_pasilla"]

    def normalizar_tipo_cafe(tipo: str) -> str:
        """Normaliza el tipo de café a clave interna."""
        tipo = tipo.strip().upper()
        if tipo == "CPS" or "CPS" in tipo or "PERGAMINO" in tipo:
            return "ingreso_cps"
        elif tipo == "PASILLA" or "PASILLA" in tipo:
            return "ingreso_pasilla"
        return tipo.lower().replace(" ", "_")

    @router.callback_query(F.data == "menu_importar")
    async def menu_importar(callback: types.CallbackQuery, state: FSMContext):
        """Inicia el flujo de importación."""
        await callback.answer()
        await state.clear()
        await state.set_state(ImportExcelState.esperando_archivo)

        keyboard = types.InlineKeyboardMarkup(
            inline_keyboard=[
                [types.InlineKeyboardButton(
                    text="📋 Descargar plantilla vacía",
                    callback_data="importar:descargar_plantilla",
                )],
            ]
        )
        keyboard = agregar_boton_menu(keyboard)

        await callback.message.answer(
            "📥 <b>Importar Excel</b>\n\n"
            "Enviame el archivo <b>.xlsx</b> con los datos que querés importar.\n\n"
            "📌 <b>Formato esperado:</b>\n"
            "El archivo debe tener una o más de estas hojas:\n"
            "• <b>Fincas</b> — nombre, region, departamento\n"
            "• <b>Lotes</b> — finca_nombre, nombre, area_hectareas, num_arboles, variedad, fecha_siembra\n"
            "• <b>Ingresos</b> — finca_nombre, tipo, fecha, cantidad, valor_total\n"
            "• <b>Costos_MO</b> — finca_nombre, lote_nombre, categoria, fecha, labor, cantidad, valor_unitario, valor_total\n"
            "• <b>Costos_Insumos</b> — finca_nombre, lote_nombre, categoria, fecha, producto, cantidad, unidad, valor_unitario, valor_total\n\n"
            "🔴 <b>Importante:</b> Las fincas deben existir o crearse primero.\n\n"
            "❌ Enviá /cancelar para salir.",
            parse_mode="HTML",
            reply_markup=keyboard,
        )

    @router.callback_query(F.data == "importar:descargar_plantilla")
    async def descargar_plantilla(callback: types.CallbackQuery):
        """Genera y envía una plantilla Excel vacía basada en el template real con instrucciones."""
        await callback.answer()

        try:
            from excel_manager import ExcelManager
            from config import EXCEL_TEMPLATE, EXPORTS_DIR

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"plantilla_importar_{timestamp}.xlsx"
            output_path = os.path.join(EXPORTS_DIR, output_filename)

            # Usar ExcelManager para generar plantilla desde el template real
            manager = ExcelManager(EXCEL_TEMPLATE)
            manager.generar_plantilla_vacia(output_path)

            await callback.message.answer_document(
                types.FSInputFile(output_path, filename="Plantilla_Importacion_Costos.xlsx"),
                caption="📋 <b>Plantilla para importar</b> ✅\n\n"
                        "📌 <b>Instrucciones:</b>\n"
                        "1. Abrí el archivo (tiene datos de ejemplo como guía)\n"
                        "2. Reemplazá los datos de ejemplo con tus datos reales\n"
                        "3. Guardá el archivo\n"
                        "4. Enviame el archivo completado\n\n"
                        "📄 La hoja <b>NOTAS</b> tiene instrucciones detalladas de cada hoja.",
                parse_mode="HTML",
            )

            # Limpiar archivo temporal
            try:
                os.remove(output_path)
            except Exception:
                pass

        except Exception as e:
            logger.error(f"Error al generar plantilla de importación: {e}", exc_info=True)
            await callback.message.answer(
                "❌ <b>Error al generar la plantilla.</b>\n\n"
                "Intentá de nuevo más tarde.",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )

    @router.message(F.document, ImportExcelState.esperando_archivo)
    async def recibir_archivo(message: types.Message, state: FSMContext):
        """Recibe y parsea el archivo Excel."""
        user_id = message.from_user.id

        # Verificar que sea .xlsx
        doc = message.document
        if not doc.file_name or not doc.file_name.lower().endswith(".xlsx"):
            await message.answer(
                "❌ <b>Formato incorrecto.</b>\n\n"
                "Solo acepto archivos <b>.xlsx</b>. Enviame un archivo válido.",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )
            return

        # Descargar archivo
        await message.answer("⏳ <b>Procesando archivo...</b>", parse_mode="HTML")

        try:
            import openpyxl

            # Descargar a archivo temporal
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name
                await message.bot.download(file=doc.file_id, destination=tmp_path)

            # Abrir con openpyxl
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            hojas_disponibles = wb.sheetnames
            logger.info(f"Archivo recibido: {doc.file_name}, hojas: {hojas_disponibles}")

            # Parsear datos por hoja
            datos_parseados = {}
            errores = []

            for hoja_nombre, columnas_esperadas in HOJAS_ESPERADAS.items():
                if hoja_nombre not in hojas_disponibles:
                    continue

                ws = wb[hoja_nombre]
                filas = list(ws.iter_rows(values_only=True))
                if not filas:
                    continue

                # Primera fila = headers
                headers = [str(h).strip().lower() if h else "" for h in filas[0]]

                # Validar headers
                columnas_faltantes = [c for c in columnas_esperadas if c not in headers]
                if columnas_faltantes:
                    errores.append(
                        f"⚠️ Hoja '{hoja_nombre}': faltan columnas: {', '.join(columnas_faltantes)}"
                    )
                    continue

                # Mapear índices
                idx_map = {}
                for col in columnas_esperadas:
                    try:
                        idx_map[col] = headers.index(col)
                    except ValueError:
                        pass

                # Parsear datos (saltar header)
                registros = []
                for i, row in enumerate(filas[1:], start=2):
                    if all(v is None for v in row):
                        continue  # Saltar filas vacías
                    registro = {}
                    es_valido = True
                    for col, idx in idx_map.items():
                        val = row[idx] if idx < len(row) else None
                        if val is not None:
                            if isinstance(val, datetime):
                                val = val.strftime("%Y-%m-%d")
                            elif hasattr(val, "strftime"):
                                val = str(val)
                            else:
                                val = str(val).strip()
                        registro[col] = val if val else ""

                    # Validaciones específicas por hoja
                    if hoja_nombre == "Fincas":
                        if not registro.get("nombre"):
                            errores.append(f"❌ Fincas fila {i}: falta el nombre")
                            es_valido = False

                    elif hoja_nombre == "Ingresos":
                        if not registro.get("finca_nombre"):
                            errores.append(f"❌ Ingresos fila {i}: falta finca_nombre")
                            es_valido = False
                        # Normalizar tipo
                        tipo_raw = registro.get("tipo", "")
                        if tipo_raw:
                            registro["tipo"] = normalizar_tipo_cafe(tipo_raw)

                    elif hoja_nombre == "Costos_MO":
                        if not registro.get("finca_nombre"):
                            errores.append(f"❌ Costos_MO fila {i}: falta finca_nombre")
                            es_valido = False
                        cat = registro.get("categoria", "")
                        if cat and cat not in CATEGORIAS_MO_VALIDAS:
                            errores.append(
                                f"❌ Costos_MO fila {i}: categoría '{cat}' no válida. "
                                f"Usar: {', '.join(CATEGORIAS_MO_VALIDAS)}"
                            )
                            es_valido = False

                    elif hoja_nombre == "Costos_Insumos":
                        if not registro.get("finca_nombre"):
                            errores.append(f"❌ Costos_Insumos fila {i}: falta finca_nombre")
                            es_valido = False
                        cat = registro.get("categoria", "")
                        if cat and cat not in CATEGORIAS_INSUMOS_VALIDAS:
                            errores.append(
                                f"❌ Costos_Insumos fila {i}: categoría '{cat}' no válida. "
                                f"Usar: {', '.join(CATEGORIAS_INSUMOS_VALIDAS)}"
                            )
                            es_valido = False

                    if es_valido:
                        registros.append(registro)

                if registros:
                    datos_parseados[hoja_nombre] = registros

            wb.close()
            os.unlink(tmp_path)

            # Mostrar errores si hay
            if errores:
                msg_error = "⚠️ <b>Errores encontrados:</b>\n\n"
                for err in errores[:10]:  # Mostrar max 10
                    msg_error += f"{err}\n"
                if len(errores) > 10:
                    msg_error += f"... y {len(errores) - 10} más\n"
                msg_error += "\nCorregí el archivo y enviálo de nuevo."
                await message.answer(msg_error, parse_mode="HTML", reply_markup=boton_menu())
                return

            if not datos_parseados:
                await message.answer(
                    "❌ <b>No se encontraron datos válidos.</b>\n\n"
                    "El archivo debe tener al menos una hoja con datos: "
                    "Fincas, Lotes, Ingresos, Costos_MO o Costos_Insumos.\n\n"
                    "Enviá el archivo corregido o /cancelar.",
                    parse_mode="HTML",
                    reply_markup=boton_menu(),
                )
                return

            # Guardar datos parseados en estado
            await state.update_data(datos_importados=datos_parseados)
            await state.set_state(ImportExcelState.preview_mostrado)

            # Mostrar preview
            preview = "📋 <b>Vista previa de datos a importar:</b>\n\n"

            for hoja, registros in datos_parseados.items():
                preview += f"📄 <b>{hoja}</b>: {len(registros)} registro(s)\n"
                # Mostrar primeros 3 registros como ejemplo
                for j, reg in enumerate(registros[:3]):
                    preview += f"  {j+1}. "
                    if hoja == "Fincas":
                        preview += f"{reg.get('nombre', '?')}"
                        if reg.get("region"):
                            preview += f" — {reg['region']}"
                    elif hoja == "Lotes":
                        preview += f"{reg.get('nombre', '?')} (Finca: {reg.get('finca_nombre', '?')})"
                    elif hoja == "Ingresos":
                        preview += f"{reg.get('tipo', '?')} — ${reg.get('valor_total', '0')} ({reg.get('fecha', '?')})"
                    elif hoja == "Costos_MO":
                        preview += f"{reg.get('categoria', '?')} — ${reg.get('valor_total', '0')} ({reg.get('labor', '?')})"
                    elif hoja == "Costos_Insumos":
                        preview += f"{reg.get('categoria', '?')} — {reg.get('producto', '?')} — ${reg.get('valor_total', '0')}"
                    preview += "\n"

                if len(registros) > 3:
                    preview += f"  ... y {len(registros) - 3} más\n"
                preview += "\n"

            keyboard = types.InlineKeyboardMarkup(
                inline_keyboard=[
                    [
                        types.InlineKeyboardButton(text="✅ Confirmar importación", callback_data="importar:confirmar"),
                        types.InlineKeyboardButton(text="❌ Cancelar", callback_data="importar:cancelar"),
                    ],
                ]
            )
            keyboard = agregar_boton_menu(keyboard)

            await message.answer(
                preview + "¿Querés importar estos datos?",
                parse_mode="HTML",
                reply_markup=keyboard,
            )

        except Exception as e:
            logger.error(f"Error al procesar archivo: {e}", exc_info=True)
            # Limpiar temp si existe
            try:
                os.unlink(tmp_path)
            except Exception:
                pass
            await message.answer(
                "❌ <b>Error al procesar el archivo.</b>\n\n"
                f"Detalle: {str(e)[:200]}\n\n"
                "Asegurate de que sea un archivo .xlsx válido y enviálo de nuevo.",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )

    @router.callback_query(F.data == "importar:cancelar", ImportExcelState.preview_mostrado)
    async def cancelar_importacion(callback: types.CallbackQuery, state: FSMContext):
        """Cancela la importación."""
        await callback.answer()
        await state.clear()
        await callback.message.edit_text(
            "✅ <b>Importación cancelada.</b>\n\n"
            "Ningún dato fue modificado.",
            parse_mode="HTML",
            reply_markup=boton_menu(),
        )

    @router.callback_query(F.data == "importar:confirmar", ImportExcelState.preview_mostrado)
    async def confirmar_importacion(callback: types.CallbackQuery, state: FSMContext):
        """Confirma y ejecuta la importación."""
        await callback.answer()
        await state.set_state(ImportExcelState.confirmado)
        await callback.message.edit_text(
            "⏳ <b>Importando datos...</b>",
            parse_mode="HTML",
        )

        data = await state.get_data()
        datos_importados = data.get("datos_importados", {})
        user_id = callback.from_user.id

        try:
            resumen = {"fincas": 0, "lotes": 0, "ingresos": 0, "costos_mo": 0, "costos_insumos": 0}
            errores_import = []

            # Crear mapa de finca_nombre → finca_id
            fincas_existentes = {f["nombre"]: f["id"] for f in db.get_fincas(user_id)}
            fincas_creadas = {}

            # 1️⃣ Importar Fincas
            if "Fincas" in datos_importados:
                for reg in datos_importados["Fincas"]:
                    nombre = reg.get("nombre", "").strip()
                    if not nombre:
                        continue
                    if nombre in fincas_existentes:
                        fincas_creadas[nombre] = fincas_existentes[nombre]
                    elif nombre in fincas_creadas:
                        continue
                    else:
                        try:
                            fid = db.create_finca(
                                user_id,
                                nombre,
                                reg.get("region", ""),
                                reg.get("departamento", ""),
                            )
                            fincas_creadas[nombre] = fid
                            resumen["fincas"] += 1
                        except Exception as e:
                            errores_import.append(f"Finca '{nombre}': {e}")

            # Combinar fincas existentes + recién creadas
            todas_fincas = {**fincas_existentes, **fincas_creadas}

            # 2️⃣ Importar Lotes
            if "Lotes" in datos_importados:
                for reg in datos_importados["Lotes"]:
                    fn = reg.get("finca_nombre", "").strip()
                    if fn not in todas_fincas:
                        errores_import.append(f"Lote '{reg.get('nombre', '?')}': finca '{fn}' no encontrada")
                        continue
                    try:
                        db.create_lote(
                            todas_fincas[fn],
                            reg.get("nombre", ""),
                            float(reg.get("area_hectareas", 0)) if reg.get("area_hectareas") else 0,
                            int(reg.get("num_arboles", 0)) if reg.get("num_arboles") else 0,
                            reg.get("variedad", ""),
                            reg.get("fecha_siembra", ""),
                        )
                        resumen["lotes"] += 1
                    except Exception as e:
                        errores_import.append(f"Lote '{reg.get('nombre', '?')}': {e}")

            # Obtener lotes de todas las fincas para mapear nombres
            lotes_por_finca = {}
            for fid in todas_fincas.values():
                for l in db.get_lotes(fid):
                    fn_key = next((n for n, id_ in todas_fincas.items() if id_ == fid), str(fid))
                    if fn_key not in lotes_por_finca:
                        lotes_por_finca[fn_key] = {}
                    lotes_por_finca[fn_key][l["nombre"]] = l["id"]

            # 3️⃣ Importar Ingresos
            if "Ingresos" in datos_importados:
                for reg in datos_importados["Ingresos"]:
                    fn = reg.get("finca_nombre", "").strip()
                    if fn not in todas_fincas:
                        errores_import.append(f"Ingreso: finca '{fn}' no encontrada")
                        continue
                    try:
                        db.insert_transaccion(
                            todas_fincas[fn],
                            reg.get("tipo", "ingreso_cps"),
                            reg.get("fecha", datetime.now().strftime("%Y-%m-%d")),
                            labor="Venta de café",
                            cantidad=float(reg.get("cantidad", 0)) if reg.get("cantidad") else 0,
                            valor_total=float(reg.get("valor_total", 0)) if reg.get("valor_total") else 0,
                        )
                        resumen["ingresos"] += 1
                    except Exception as e:
                        errores_import.append(f"Ingreso: {e}")

            # 4️⃣ Importar Costos MO
            if "Costos_MO" in datos_importados:
                for reg in datos_importados["Costos_MO"]:
                    fn = reg.get("finca_nombre", "").strip()
                    if fn not in todas_fincas:
                        errores_import.append(f"Costo MO: finca '{fn}' no encontrada")
                        continue
                    lote_id = 0
                    ln = reg.get("lote_nombre", "").strip()
                    if ln and fn in lotes_por_finca and ln in lotes_por_finca[fn]:
                        lote_id = lotes_por_finca[fn][ln]
                    try:
                        db.insert_transaccion(
                            todas_fincas[fn],
                            reg.get("categoria", ""),
                            reg.get("fecha", datetime.now().strftime("%Y-%m-%d")),
                            labor=reg.get("labor", ""),
                            cantidad=float(reg.get("cantidad", 0)) if reg.get("cantidad") else 0,
                            valor_unitario=float(reg.get("valor_unitario", 0)) if reg.get("valor_unitario") else 0,
                            valor_total=float(reg.get("valor_total", 0)) if reg.get("valor_total") else 0,
                            lote_id=lote_id,
                        )
                        resumen["costos_mo"] += 1
                    except Exception as e:
                        errores_import.append(f"Costo MO: {e}")

            # 5️⃣ Importar Costos Insumos
            if "Costos_Insumos" in datos_importados:
                for reg in datos_importados["Costos_Insumos"]:
                    fn = reg.get("finca_nombre", "").strip()
                    if fn not in todas_fincas:
                        errores_import.append(f"Costo Insumo: finca '{fn}' no encontrada")
                        continue
                    lote_id = 0
                    ln = reg.get("lote_nombre", "").strip()
                    if ln and fn in lotes_por_finca and ln in lotes_por_finca[fn]:
                        lote_id = lotes_por_finca[fn][ln]
                    try:
                        db.insert_transaccion(
                            todas_fincas[fn],
                            reg.get("categoria", ""),
                            reg.get("fecha", datetime.now().strftime("%Y-%m-%d")),
                            producto=reg.get("producto", ""),
                            cantidad=float(reg.get("cantidad", 0)) if reg.get("cantidad") else 0,
                            unidad=reg.get("unidad", ""),
                            valor_unitario=float(reg.get("valor_unitario", 0)) if reg.get("valor_unitario") else 0,
                            valor_total=float(reg.get("valor_total", 0)) if reg.get("valor_total") else 0,
                            lote_id=lote_id,
                        )
                        resumen["costos_insumos"] += 1
                    except Exception as e:
                        errores_import.append(f"Costo Insumo: {e}")

            # Construir mensaje de resultado
            total_importados = sum(resumen.values())
            resultado = f"✅ <b>Importación completada</b>\n\n"
            resultado += f"📊 <b>Resumen:</b>\n"
            resultado += f"  🏠 Fincas: {resumen['fincas']} creadas\n"
            resultado += f"  🌱 Lotes: {resumen['lotes']}\n"
            resultado += f"  💰 Ingresos: {resumen['ingresos']}\n"
            resultado += f"  🔧 Costos MO: {resumen['costos_mo']}\n"
            resultado += f"  📦 Costos Insumos: {resumen['costos_insumos']}\n"
            resultado += f"\n📌 <b>Total:</b> {total_importados} registro(s) importado(s)\n"

            if errores_import:
                resultado += f"\n⚠️ <b>Advertencias/Errores:</b>\n"
                for err in errores_import[:5]:
                    resultado += f"  • {err}\n"
                if len(errores_import) > 5:
                    resultado += f"  ... y {len(errores_import) - 5} más\n"

            await callback.message.edit_text(resultado, parse_mode="HTML", reply_markup=boton_menu())

        except Exception as e:
            logger.error(f"Error en importación: {e}", exc_info=True)
            await callback.message.edit_text(
                f"❌ <b>Error durante la importación.</b>\n\n{str(e)[:300]}",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )

        finally:
            await state.clear()

    return router
