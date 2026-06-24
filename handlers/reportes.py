"""
Handler de /resumen - Reportes y exportación de Excel.
"""
import os
import logging
from datetime import datetime, timedelta
from aiogram import Router, types, F
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from database import Database
from config import CATEGORIAS, EXPORTS_DIR, EXCEL_TEMPLATE, ADMIN_IDS
from utils import boton_menu, agregar_boton_menu

logger = logging.getLogger(__name__)


class FiltrarStates(StatesGroup):
    esperando_fecha_inicio = State()
    esperando_fecha_fin = State()


def _split_texto(texto: str, max_len: int = 4096) -> list:
    """Divide un texto largo en partes."""
    if len(texto) <= max_len:
        return [texto]

    partes = []
    while len(texto) > max_len:
        # Buscar el último salto de línea antes del límite
        corte = texto.rfind("\n", 0, max_len)
        if corte == -1:
            corte = max_len
        partes.append(texto[:corte])
        texto = texto[corte:].strip()
    if texto:
        partes.append(texto)
    return partes


async def mostrar_resumen(db: Database, send_func, finca_id: int, finca_nombre: str):
    """Muestra el resumen detallado de una finca."""
    resumen = db.get_resumen_finca(finca_id)
    if not resumen:
        await send_func("❌ <b>Error al obtener el resumen.</b>", parse_mode="HTML")
        return

    total_ingresos = resumen.get("ingresos", 0)
    total_egresos = resumen.get("egresos", 0)
    margen = resumen.get("margen", 0)
    area = resumen.get("area_total", 0)
    costo_ha = resumen.get("costo_por_hectarea", 0)

    texto = (
        f"📊 <b>Resumen — {finca_nombre}</b>\n\n"
        f"📐 <b>Área total:</b> {area:.2f} ha\n\n"
        f"💰 <b>Ingresos Totales:</b> ${total_ingresos:,.0f}\n"
        f"📉 <b>Egresos Totales:</b> ${total_egresos:,.0f}\n"
    )

    if margen >= 0:
        texto += f"✅ <b>Margen:</b> ${margen:,.0f}\n"
    else:
        texto += f"❌ <b>Margen:</b> -${abs(margen):,.0f} (pérdida)\n"

    texto += f"📊 <b>Costo por hectárea:</b> ${costo_ha:,.0f}\n\n"

    # Egresos por categoría
    if resumen["egresos_por_categoria"]:
        texto += "<b>Egresos por categoría:</b>\n"
        for cat, total in sorted(resumen["egresos_por_categoria"].items(), key=lambda x: x[1], reverse=True):
            nombre_cat = CATEGORIAS.get(cat, {}).get("nombre", cat)
            porcentaje = (total / total_egresos * 100) if total_egresos > 0 else 0
            texto += f"  • {nombre_cat}: ${total:,.0f} ({porcentaje:.1f}%)\n"

    # Ingresos por tipo
    if resumen["ingresos_por_tipo"]:
        texto += "\n<b>Ingresos por tipo:</b>\n"
        for cat, total in resumen["ingresos_por_tipo"].items():
            nombre_cat = CATEGORIAS.get(cat, {}).get("nombre", cat)
            texto += f"  • {nombre_cat}: ${total:,.0f}\n"

    texto += "\n<b>Acciones:</b>"

    keyboard = types.InlineKeyboardMarkup(
        inline_keyboard=[
            [types.InlineKeyboardButton(
                text="📊 Generar Excel",
                callback_data=f"generar_excel:{finca_id}",
            )],
            [types.InlineKeyboardButton(
                text="📄 Exportar PDF",
                callback_data=f"resumen_pdf:{finca_id}",
            )],
        ]
    )
    keyboard = agregar_boton_menu(keyboard)

    # Manejar textos largos
    partes = _split_texto(texto)
    for i, parte in enumerate(partes):
        if i == len(partes) - 1:
            await send_func(parte, parse_mode="HTML", reply_markup=keyboard)
        else:
            await send_func(parte, parse_mode="HTML")


def get_reportes_router(db: Database) -> Router:
    router = Router()

    @router.message(Command("resumen"))
    @router.callback_query(F.data == "menu_resumen")
    async def cmd_resumen(event: types.Message | types.CallbackQuery, state: FSMContext):
        """Muestra el resumen de la finca."""
        await state.clear()
        user_id = event.from_user.id

        if isinstance(event, types.CallbackQuery):
            await event.answer()
            message = event.message
            send = message.answer
        else:
            message = event
            send = message.answer

        if not db.is_approved(user_id):
            await send("⏳ <b>No tienes acceso.</b> Usa /start para solicitar aprobación.", parse_mode="HTML", reply_markup=boton_menu())
            return

        try:
            fincas = db.get_fincas(user_id)
            if not fincas:
                await send(
                    "❌ <b>No tienes fincas registradas.</b>\n\n"
                    "Primero crea una finca con /fincas 🗺️",
                    parse_mode="HTML",
                    reply_markup=boton_menu(),
                )
                return

            if len(fincas) == 1:
                await mostrar_resumen(db, send, fincas[0]["id"], fincas[0]["nombre"])
                return

            # Varias fincas
            keyboard = types.InlineKeyboardMarkup(
                inline_keyboard=[
                    [types.InlineKeyboardButton(
                        text=f"🏠 {f['nombre']}",
                        callback_data=f"resumen_finca:{f['id']}",
                    )]
                    for f in fincas
                ]
                + [
                    [types.InlineKeyboardButton(text="🔙 Volver", callback_data="volver_menu")],
                ]
            )

            await send(
                "📊 <b>Resumen de Finca</b>\n\nSelecciona la finca:",
                parse_mode="HTML",
                reply_markup=keyboard,
            )

        except Exception as e:
            logger.error(f"Error en /resumen: {e}", exc_info=True)
            await send("❌ <b>Error al generar resumen.</b>", parse_mode="HTML", reply_markup=boton_menu())

    @router.callback_query(F.data.startswith("resumen_finca:"))
    async def seleccionar_finca_resumen(callback: types.CallbackQuery):
        await callback.answer()
        user_id = callback.from_user.id
        finca_id = int(callback.data.split(":")[1])
        finca = db.get_finca_by_id(finca_id)
        if not finca:
            await callback.message.edit_text("❌ <b>Finca no encontrada.</b>", parse_mode="HTML")
            return
        if finca["user_id"] != user_id:
            await callback.message.edit_text("❌ <b>Esta finca no te pertenece.</b>", parse_mode="HTML")
            return

        await mostrar_resumen(db, callback.message.edit_text, finca_id, finca["nombre"])

    @router.callback_query(F.data == "menu_excel")
    @router.callback_query(F.data.startswith("generar_excel:"))
    async def cmd_generar_excel(callback: types.CallbackQuery):
        """Genera y envía el Excel."""
        await callback.answer()

        if callback.data == "menu_excel":
            # Preguntar qué finca
            user_id = callback.from_user.id
            fincas = db.get_fincas(user_id)
            if not fincas:
                # ✅ Generar plantilla vacía usando ExcelManager
                try:
                    await callback.message.edit_text(
                        "⏳ <b>Generando plantilla Excel vacía...</b>",
                        parse_mode="HTML",
                    )

                    from excel_manager import ExcelManager
                    manager = ExcelManager(EXCEL_TEMPLATE)

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    output_filename = f"plantilla_vacia_{timestamp}.xlsx"
                    output_path = os.path.join(EXPORTS_DIR, output_filename)

                    # Usar el método del manager
                    manager.generar_plantilla_vacia(output_path)

                    # Enviar archivo
                    with open(output_path, "rb") as f:
                        await callback.message.answer_document(
                            types.FSInputFile(output_path, filename="Plantilla_Vacia.xlsx"),
                            caption="📋 <b>Plantilla Excel generada</b> ☕\n\n"
                                    "No tenés fincas registradas, pero acá tenés el formato "
                                    "del Excel con datos de ejemplo como guía.\n"
                                    "Podés llenarlo manualmente o crear una finca "
                                    "con /fincas y volver a exportar.",
                            parse_mode="HTML",
                            reply_markup=boton_menu(),
                        )

                    # Limpiar archivo temporal
                    try:
                        os.remove(output_path)
                    except Exception:
                        pass

                except FileNotFoundError as e:
                    logger.error(f"Template no encontrado: {e}")
                    await callback.message.edit_text(
                        "❌ <b>Error:</b> El template Excel no está disponible.\n\n"
                        "Contacta al administrador.",
                        parse_mode="HTML",
                        reply_markup=boton_menu(),
                    )
                except Exception as e:
                    logger.error(f"Error al generar plantilla vacía: {e}", exc_info=True)
                    await callback.message.edit_text(
                        "❌ <b>Error al generar la plantilla.</b> Intenta de nuevo más tarde.",
                        parse_mode="HTML",
                        reply_markup=boton_menu(),
                    )
                return

            if len(fincas) == 1:
                finca_id = fincas[0]["id"]
            else:
                keyboard = types.InlineKeyboardMarkup(
                    inline_keyboard=[
                        [types.InlineKeyboardButton(
                            text=f"🏠 {f['nombre']}",
                            callback_data=f"generar_excel:{f['id']}",
                        )]
                        for f in fincas
                    ]
                    + [
                        [types.InlineKeyboardButton(text="🔙 Volver", callback_data="volver_menu")],
                    ]
                )
                await callback.message.edit_text(
                    "📊 <b>Generar Excel</b>\n\nSelecciona la finca:",
                    parse_mode="HTML",
                    reply_markup=keyboard,
                )
                return
        else:
            finca_id = int(callback.data.split(":")[1])

        try:
            await callback.message.edit_text(
                "⏳ <b>Generando Excel...</b> Esto puede tomar unos segundos.",
                parse_mode="HTML",
            )

            # Generar Excel
            from excel_manager import ExcelManager
            manager = ExcelManager(EXCEL_TEMPLATE)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"costos_finca_{finca_id}_{timestamp}.xlsx"
            output_path = os.path.join(EXPORTS_DIR, output_filename)

            manager.generar_excel(finca_id, db, output_path)

            # Enviar archivo
            with open(output_path, "rb") as f:
                await callback.message.answer_document(
                    types.FSInputFile(output_path, filename=f"Costos_Produccion_{finca_id}.xlsx"),
                    caption="📊 <b>Excel de Costos de Producción generado</b> ☕\n\n"
                            "Las fórmulas de Resultados y Gráficos se calcularán "
                            "al abrir el archivo en Excel o LibreOffice.",
                    parse_mode="HTML",
                )

            # Limpiar archivo temporal
            try:
                os.remove(output_path)
            except Exception:
                pass

        except FileNotFoundError as e:
            logger.error(f"Template no encontrado: {e}")
            await callback.message.edit_text(
                "❌ <b>Error:</b> El template Excel no está disponible.\n\n"
                "Contacta al administrador.",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )
        except Exception as e:
            logger.error(f"Error al generar Excel: {e}", exc_info=True)
            await callback.message.edit_text(
                "❌ <b>Error al generar el Excel.</b> Intenta de nuevo más tarde.",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )

    @router.callback_query(F.data.startswith("resumen_pdf:"))
    async def cmd_exportar_pdf(callback: types.CallbackQuery):
        """Exporta el resumen a PDF."""
        await callback.answer()
        user_id = callback.from_user.id
        finca_id = int(callback.data.split(":")[1])
        finca = db.get_finca_by_id(finca_id)
        if not finca or finca["user_id"] != user_id:
            await callback.message.edit_text(
                "❌ <b>Finca no encontrada o no te pertenece.</b>",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )
            return

        try:
            await callback.message.edit_text(
                "⏳ <b>Generando PDF del resumen ejecutivo...</b>",
                parse_mode="HTML",
            )

            resumen = db.get_resumen_finca(finca_id)
            if not resumen or resumen.get("area_total", 0) == 0:
                await callback.message.edit_text(
                    "⚠️ <b>No hay datos suficientes para generar el PDF.</b>",
                    parse_mode="HTML",
                    reply_markup=boton_menu(),
                )
                return

            from fpdf import FPDF
            pdf = FPDF()
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 16)
            pdf.set_text_color(31, 78, 121)
            pdf.cell(0, 10, f"Resumen - {finca['nombre']}", ln=True, align="C")
            pdf.ln(10)

            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(0, 8, "RESUMEN FINANCIERO", ln=True)
            pdf.set_font("Helvetica", "", 10)

            total_ingresos = resumen.get("ingresos", 0)
            total_egresos = resumen.get("egresos", 0)
            margen = resumen.get("margen", 0)

            pdf.cell(0, 7, f"  Area Total: {resumen.get('area_total', 0):.2f} ha", ln=True)
            pdf.cell(0, 7, f"  Ingresos Totales: ${total_ingresos:,.0f}", ln=True)
            pdf.cell(0, 7, f"  Egresos Totales: ${total_egresos:,.0f}", ln=True)
            pdf.cell(0, 7, f"  Margen: ${margen:,.0f}", ln=True)
            pdf.cell(0, 7, f"  Costo por ha: ${resumen.get('costo_por_hectarea', 0):,.0f}", ln=True)

            pdf.ln(5)
            pdf.set_font("Helvetica", "B", 12)
            pdf.cell(0, 8, "EGRESOS POR CATEGORIA", ln=True)
            pdf.set_font("Helvetica", "", 10)

            from config import CATEGORIAS as CAT_MAP
            for cat, total in sorted(resumen.get("egresos_por_categoria", {}).items(), key=lambda x: x[1], reverse=True):
                nombre_cat = CAT_MAP.get(cat, {}).get("nombre", cat)
                pdf.cell(0, 7, f"  {nombre_cat}: ${total:,.0f}", ln=True)

            from config import EXPORTS_DIR
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"resumen_finca_{finca_id}_{timestamp}.pdf"
            output_path = os.path.join(EXPORTS_DIR, output_filename)
            pdf.output(output_path)

            with open(output_path, "rb") as f:
                await callback.message.answer_document(
                    types.FSInputFile(output_path, filename=f"Resumen_{finca['nombre']}.pdf"),
                    caption=f"📄 <b>Resumen PDF generado</b> ☕\n\n"
                            f"Resumen financiero de {finca['nombre']}.",
                    parse_mode="HTML",
                )

            try:
                os.remove(output_path)
            except Exception:
                pass

        except ImportError:
            await callback.message.edit_text(
                "⚠️ <b>FPDF2 no está instalado.</b>\n\n"
                "Instalalo con: pip install fpdf2\n"
                "Mientras tanto, usá la exportación a Excel.",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )
        except Exception as e:
            logger.error(f"Error al generar PDF: {e}", exc_info=True)
            await callback.message.edit_text(
                "❌ <b>Error al generar el PDF.</b> Intenta de nuevo más tarde.",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )

    # ─── FILTRAR POR PERÍODO ───

    @router.message(Command("filtrar"))
    @router.callback_query(F.data == "menu_filtrar")
    async def cmd_filtrar(event: types.Message | types.CallbackQuery, state: FSMContext):
        """Muestra el menú de filtrado por período."""
        await state.clear()
        user_id = event.from_user.id

        if isinstance(event, types.CallbackQuery):
            await event.answer()
            message = event.message
            send = message.answer
        else:
            message = event
            send = message.answer

        if not db.is_approved(user_id):
            await send("⏳ <b>No tienes acceso.</b> Usa /start para solicitar aprobación.", parse_mode="HTML", reply_markup=boton_menu())
            return

        fincas = db.get_fincas(user_id)
        if not fincas:
            await send(
                "❌ <b>No tienes fincas registradas.</b>\n\nPrimero crea una finca con /fincas 🗺️",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )
            return

        if len(fincas) == 1:
            await mostrar_menu_filtrar(send, fincas[0], db)
        else:
            keyboard = types.InlineKeyboardMarkup(
                inline_keyboard=[
                    [types.InlineKeyboardButton(
                        text=f"🏠 {f['nombre']}",
                        callback_data=f"filtrar_seleccion:{f['id']}",
                    )]
                    for f in fincas
                ] + [
                    [types.InlineKeyboardButton(text="🔙 Volver", callback_data="volver_menu")],
                ]
            )
            await send(
                "📊 <b>Filtrar por período</b>\n\nSelecciona la finca:",
                parse_mode="HTML",
                reply_markup=keyboard,
            )

    async def mostrar_menu_filtrar(send_func, finca: dict, db_instance):
        """Muestra el menú de opciones de filtrado."""
        ahora = datetime.now()
        año_actual = ahora.year
        mes_actual = ahora.month
        _, semana_actual, _ = ahora.isocalendar()

        keyboard = types.InlineKeyboardMarkup(
            inline_keyboard=[
                [types.InlineKeyboardButton(
                    text="📅 Esta semana",
                    callback_data=f"filtrar:semana:{finca['id']}:{año_actual}:{semana_actual}",
                )],
                [types.InlineKeyboardButton(
                    text="📅 Este mes",
                    callback_data=f"filtrar:mes:{finca['id']}:{año_actual}:{mes_actual}",
                )],
                [types.InlineKeyboardButton(
                    text=f"📅 Año {año_actual}",
                    callback_data=f"filtrar:anio:{finca['id']}:{año_actual}",
                )],
                [types.InlineKeyboardButton(
                    text=f"📅 Año {año_actual - 1}",
                    callback_data=f"filtrar:anio:{finca['id']}:{año_actual - 1}",
                )],
                [types.InlineKeyboardButton(
                    text="📅 Personalizado",
                    callback_data=f"filtrar:personalizado:{finca['id']}",
                )],
            ]
        )
        keyboard = agregar_boton_menu(keyboard)

        await send_func(
            f"📊 <b>Filtrar por período</b>\n\n"
            f"🏠 Finca: {finca['nombre']}\n\n"
            f"¿Qué período querés ver?",
            parse_mode="HTML",
            reply_markup=keyboard,
        )

    @router.callback_query(F.data.startswith("filtrar_seleccion:"))
    async def seleccionar_finca_filtrar(callback: types.CallbackQuery):
        await callback.answer()
        user_id = callback.from_user.id
        finca_id = int(callback.data.split(":")[1])
        finca = db.get_finca_by_id(finca_id)
        if not finca or finca["user_id"] != user_id:
            await callback.message.edit_text("❌ <b>Finca no encontrada.</b>", parse_mode="HTML")
            return
        await mostrar_menu_filtrar(callback.message.edit_text, finca, db)

    @router.callback_query(F.data.startswith("filtrar:"))
    async def ejecutar_filtro(callback: types.CallbackQuery, state: FSMContext):
        """Ejecuta el filtro seleccionado y muestra el resumen del período."""
        await callback.answer()
        user_id = callback.from_user.id
        parts = callback.data.split(":")
        tipo = parts[1]
        finca_id = int(parts[2])

        finca = db.get_finca_by_id(finca_id)
        if not finca or finca["user_id"] != user_id:
            await callback.message.edit_text("❌ <b>Finca no encontrada.</b>", parse_mode="HTML")
            return

        if tipo == "semana":
            año = int(parts[3])
            semana = int(parts[4])
            resumen = db.get_resumen_semanal(finca_id, año, semana)
            etiqueta = f"Sem {semana} {año}"
        elif tipo == "mes":
            año = int(parts[3])
            mes = int(parts[4])
            resumen = db.get_resumen_mensual(finca_id, año, mes)
            meses_nombre = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
                           "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
            etiqueta = f"{meses_nombre[mes - 1]} {año}"
        elif tipo == "anio":
            año = int(parts[3])
            resumen = db.get_resumen_anual(finca_id, año)
            etiqueta = f"Año {año}"
        elif tipo == "personalizado":
            # Pedir fecha inicio
            await state.set_state(FiltrarStates.esperando_fecha_inicio)
            await state.update_data(finca_id=finca_id, finca_nombre=finca["nombre"])
            await callback.message.edit_text(
                "📅 <b>Filtrar por período personalizado</b>\n\n"
                "Enviame la <b>fecha de inicio</b> en formato DD/MM/AAAA:\n\n"
                "Ejemplo: <code>01/01/2025</code>",
                parse_mode="HTML",
                reply_markup=agregar_boton_menu(types.InlineKeyboardMarkup(inline_keyboard=[
                    [types.InlineKeyboardButton(text="❌ Cancelar", callback_data="cancelar_operacion")],
                ])),
            )
            return
        else:
            await callback.message.edit_text("❌ <b>Opción no válida.</b>", parse_mode="HTML")
            return

        await mostrar_resumen_periodo(callback.message.edit_text, finca, resumen, etiqueta, db)

    @router.message(FiltrarStates.esperando_fecha_inicio)
    async def recibir_fecha_inicio(message: types.Message, state: FSMContext):
        """Recibe la fecha de inicio para filtro personalizado."""
        texto = message.text.strip()
        data = await state.get_data()
        finca_id = data["finca_id"]
        finca_nombre = data["finca_nombre"]

        # Intentar parsear la fecha
        fecha_inicio = None
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                fecha_inicio = datetime.strptime(texto, fmt).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

        if not fecha_inicio:
            await message.answer(
                "❌ <b>Formato de fecha no válido.</b>\n\n"
                "Usá el formato DD/MM/AAAA, por ejemplo: <code>01/01/2025</code>\n\n"
                "Intentá de nuevo:",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )
            return

        await state.update_data(fecha_inicio=fecha_inicio)
        await state.set_state(FiltrarStates.esperando_fecha_fin)
        await message.answer(
            f"✅ Fecha de inicio: <b>{texto}</b>\n\n"
            "Ahora enviame la <b>fecha de fin</b> en formato DD/MM/AAAA:\n\n"
            "Ejemplo: <code>31/12/2025</code>",
            parse_mode="HTML",
            reply_markup=agregar_boton_menu(types.InlineKeyboardMarkup(inline_keyboard=[
                [types.InlineKeyboardButton(text="❌ Cancelar", callback_data="cancelar_operacion")],
            ])),
        )

    @router.message(FiltrarStates.esperando_fecha_fin)
    async def recibir_fecha_fin(message: types.Message, state: FSMContext):
        """Recibe la fecha de fin y muestra el resumen del período."""
        texto = message.text.strip()
        data = await state.get_data()
        finca_id = data["finca_id"]
        finca_nombre = data["finca_nombre"]
        fecha_inicio = data["fecha_inicio"]

        # Intentar parsear la fecha de fin
        fecha_fin = None
        for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"]:
            try:
                fecha_fin = datetime.strptime(texto, fmt).strftime("%Y-%m-%d")
                break
            except ValueError:
                continue

        if not fecha_fin:
            await message.answer(
                "❌ <b>Formato de fecha no válido.</b>\n\n"
                "Usá el formato DD/MM/AAAA, por ejemplo: <code>31/12/2025</code>\n\n"
                "Intentá de nuevo:",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )
            return

        await state.clear()

        finca = db.get_finca_by_id(finca_id)
        if not finca:
            await message.answer("❌ <b>Finca no encontrada.</b>", parse_mode="HTML")
            return

        resumen = db.get_resumen_por_periodo(finca_id, fecha_inicio, fecha_fin)
        etiqueta = f"{fecha_inicio} a {fecha_fin}"
        await mostrar_resumen_periodo(message.answer, finca, resumen, etiqueta, db)

    async def mostrar_resumen_periodo(send_func, finca: dict, resumen: dict, etiqueta: str, db_instance):
        """Muestra el resumen del período filtrado."""
        if not resumen or (resumen["ingresos"] == 0 and resumen["egresos"] == 0):
            await send_func(
                f"📭 <b>No hay datos en el período {etiqueta}</b>\n\n"
                f"No se encontraron transacciones en este período para <b>{finca['nombre']}</b>.\n\n"
                f"Probá con otro período o registrá datos con /ingresos y /costos.",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )
            return

        ingresos = resumen["ingresos"]
        egresos = resumen["egresos"]
        margen = resumen["margen"]

        texto = (
            f"📊 <b>Filtrar por período — {finca['nombre']}</b>\n\n"
            f"📅 <b>Período:</b> {etiqueta}\n\n"
            f"💰 <b>Ingresos:</b> ${ingresos:,.0f}\n"
            f"📉 <b>Egresos:</b> ${egresos:,.0f}\n"
        )

        if margen >= 0:
            texto += f"✅ <b>Margen:</b> ${margen:,.0f}\n"
        else:
            texto += f"❌ <b>Margen:</b> -${abs(margen):,.0f} (pérdida)\n"

        # Egresos por categoría
        if resumen["egresos_por_categoria"]:
            texto += "\n<b>Egresos por categoría:</b>\n"
            for cat, total in sorted(resumen["egresos_por_categoria"].items(), key=lambda x: x[1], reverse=True):
                if total > 0:
                    nombre_cat = CATEGORIAS.get(cat, {}).get("nombre", cat)
                    pct = (total / egresos * 100) if egresos > 0 else 0
                    texto += f"  • {nombre_cat}: ${total:,.0f} ({pct:.1f}%)\n"

        # Ingresos por tipo
        if resumen["ingresos_por_tipo"]:
            texto += "\n<b>Ingresos por tipo:</b>\n"
            for cat, total in resumen["ingresos_por_tipo"].items():
                if total > 0:
                    nombre_cat = CATEGORIAS.get(cat, {}).get("nombre", cat)
                    texto += f"  • {nombre_cat}: ${total:,.0f}\n"

        # Botones de acción
        keyboard = types.InlineKeyboardMarkup(
            inline_keyboard=[
                [types.InlineKeyboardButton(
                    text="📊 Generar Excel del período",
                    callback_data=f"excel_periodo:{finca['id']}:{resumen.get('fecha_inicio', '')}:{resumen.get('fecha_fin', '')}:{etiqueta}",
                )],
                [types.InlineKeyboardButton(
                    text="📅 Otro período",
                    callback_data="menu_filtrar",
                )],
            ]
        )
        keyboard = agregar_boton_menu(keyboard)

        # Manejar textos largos
        partes = _split_texto(texto)
        for i, parte in enumerate(partes):
            if i == len(partes) - 1:
                await send_func(parte, parse_mode="HTML", reply_markup=keyboard)
            else:
                await send_func(parte, parse_mode="HTML")

    @router.callback_query(F.data.startswith("excel_periodo:"))
    async def generar_excel_periodo(callback: types.CallbackQuery):
        """Genera Excel con la hoja de período incluida."""
        await callback.answer()
        user_id = callback.from_user.id
        parts = callback.data.split(":", 4)
        finca_id = int(parts[1])
        fecha_inicio = parts[2]
        fecha_fin = parts[3]
        etiqueta = parts[4] if len(parts) > 4 else ""

        finca = db.get_finca_by_id(finca_id)
        if not finca or finca["user_id"] != user_id:
            await callback.message.edit_text("❌ <b>Finca no encontrada.</b>", parse_mode="HTML")
            return

        try:
            await callback.message.edit_text(
                "⏳ <b>Generando Excel con datos del período...</b>",
                parse_mode="HTML",
            )

            from excel_manager import ExcelManager
            manager = ExcelManager(EXCEL_TEMPLATE)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"costos_periodo_{finca_id}_{timestamp}.xlsx"
            output_path = os.path.join(EXPORTS_DIR, output_filename)

            # Generar Excel con la hoja de período adicional
            manager.generar_excel(finca_id, db, output_path)

            # Agregar hoja de período al Excel generado
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            manager._llenar_hoja_periodo(wb, db, finca_id, fecha_inicio, fecha_fin, etiqueta)
            wb.save(output_path)
            wb.close()

            with open(output_path, "rb") as f:
                await callback.message.answer_document(
                    types.FSInputFile(output_path, filename=f"Costos_Periodo_{finca_id}.xlsx"),
                    caption=f"📊 <b>Excel del período generado</b> ☕\n\n"
                            f"Finca: {finca['nombre']}\n"
                            f"Período: {fecha_inicio} al {fecha_fin}\n\n"
                            f"Incluye una hoja 'Período {etiqueta}' con los datos filtrados.",
                    parse_mode="HTML",
                )

            try:
                os.remove(output_path)
            except Exception:
                pass

        except Exception as e:
            logger.error(f"Error al generar Excel de período: {e}", exc_info=True)
            await callback.message.edit_text(
                "❌ <b>Error al generar el Excel.</b> Intenta de nuevo más tarde.",
                parse_mode="HTML",
                reply_markup=boton_menu(),
            )

    return router
